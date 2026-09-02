#!/usr/bin/env python3
"""Export every table in bluedart_escalation to one .xlsx, a sheet per table.

Adds two reference sheets ahead of the data so the structure reads on its own:
  _schema         every column, type, nullability, default and key role
  _relationships  every foreign key, as from-column -> to-column

No database driver needed - psql does the reading in CSV mode.
"""
import csv
import io
import os
import re
import subprocess
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Beside this script's repo, not a hardcoded server path, so the exporter runs
# from any checkout. BLUEDART_ENV overrides it.
ENV = Path(os.environ.get("BLUEDART_ENV")
           or Path(__file__).resolve().parent.parent / ".env")
OUT = Path(sys.argv[1] if len(sys.argv) > 1 else "/tmp/bluedart_escalation.xlsx")

# Dependency order, so the sheets read the way the data is written.
TABLES = [
    ("runs",                  "id"),
    ("tickets",               "ticket_number"),
    ("awb_registry",          "awb"),
    ("clickpost_statuses",    "id"),
    ("emails_sent",           "id"),
    ("email_replies",         "id"),
    ("email_tickets",         "email_id, ticket_number"),
    ("reply_ticket_statuses", "id"),
    ("email_followups",       "email_id, sent_at"),
    ("ticket_events",         "ticket_number, occurred_at, id"),
    # A view, not a table - SELECT works the same and this is the sheet most
    # people will actually open.
    ("ticket_journey",        "ticket_number"),
]

ROLE = {
    "emails_sent":           "MAIN - one row per outbound mail",
    "email_tickets":         "LINE ITEMS - one row per ticket per mail, holds latest status",
    "email_replies":         "REPLY - one row per inbound message, append-only",
    "reply_ticket_statuses": "HISTORY - one row per ticket per reply, append-only",
    "tickets":               "SPINE - Desk ticket mirror, everything joins here",
    "runs":                  "Run ledger and watermark (replaces state.json)",
    "awb_registry":          "One keeper ticket per AWB (replaces awb_registry.json)",
    "clickpost_statuses":    "Courier snapshot per run per AWB",
    "email_followups":       "HISTORY - one row per chase sent, append-only",
    "ticket_events":         "TIMELINE - every event in a ticket's life, append-only",
    "ticket_journey":        "REPORTING VIEW - one row per ticket, the whole journey",
}

HDR_FILL = PatternFill("solid", fgColor="1F52C8")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
SUB_FILL = PatternFill("solid", fgColor="E8EEFC")
MONO = Font(name="Consolas", size=10)


def url():
    for line in ENV.read_text().splitlines():
        if line.startswith("DATABASE_URL="):
            return line.split("=", 1)[1].strip()
    raise SystemExit("DATABASE_URL not found in .env")


M = re.match(r"postgresql://([^:]+):([^@]+)@([^:]+):(\d+)/(.+)", url())
USER, PW, HOST, PORT, DB = M.groups()


def csv_query(sql):
    """Run a query and return (header, rows)."""
    r = subprocess.run(
        ["psql", "-h", HOST, "-p", PORT, "-U", USER, "-d", DB, "--csv",
         "-c", sql],
        capture_output=True, text=True,
        # Inherit the caller's environment rather than replacing it: pinning PATH
        # to /usr/bin:/bin meant psql could only ever be found on one server.
        #
        # PGTZ, not a second -c "SET TIME ZONE": psql prints "SET" for that
        # statement too, and --csv then reads it as the header row, shifting
        # every column by one. libpq applies PGTZ to the session silently.
        # It matters because timestamptz renders in the session's zone and
        # everything here happened in IST - without it every time is 5h30m out.
        env={**os.environ, "PGPASSWORD": PW, "PGTZ": "Asia/Kolkata"})
    if r.returncode:
        raise SystemExit("psql failed: " + r.stderr.strip())
    rows = list(csv.reader(io.StringIO(r.stdout)))
    return (rows[0], rows[1:]) if rows else ([], [])


# Postgres renders these as text through --csv. Excel - and Zoho Analytics
# reading the sheet - want real date cells, so they are converted back.
TS_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})[ T]"
                   r"(\d{2}):(\d{2}):(\d{2})(?:\.\d+)?"
                   r"(?:[+-]\d{2}(?::?\d{2})?)?$")
DATE_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})$")

# What Zoho Analytics is told to expect: dd-MM-yyyy HH:mm:ss. The cells hold
# real datetimes, so the value is unambiguous whatever the format shows.
DT_FMT = "DD-MM-YYYY HH:MM:SS"
D_FMT = "DD-MM-YYYY"


def as_date(v):
    """A real datetime/date for a timestamp string, else None."""
    if not isinstance(v, str) or not v:
        return None
    m = TS_RE.match(v)
    if m:
        y, mo, d, hh, mi, ss = (int(x) for x in m.groups())
        # tzinfo is dropped on purpose: the value is already IST and Excel has
        # no concept of a zone, so carrying one would only invite confusion.
        return datetime(y, mo, d, hh, mi, ss), DT_FMT
    m = DATE_RE.match(v)
    if m:
        return date(*(int(x) for x in m.groups())), D_FMT
    return None


# Filled from information_schema once it has been read: {(relation, column):
# data_type}. Counts are converted to real numbers off the DECLARED type, never
# by guessing from the digits - ticket_number and awb are text and must stay
# text, leading zeros and all.
TYPES = {}
NUMERIC = ("integer", "bigint", "smallint", "numeric", "double precision",
           "real")


def style(ws, header, rows, widths_from=None, relation=None):
    """Header styling, freeze, autofilter, sane column widths."""
    for i, h in enumerate(header, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font = HDR_FILL, HDR_FONT
        c.alignment = Alignment(vertical="center")
    for r, row in enumerate(rows, 2):
        for i, v in enumerate(row, 1):
            # Excel rejects anything over 32767 chars in one cell.
            if isinstance(v, str) and len(v) > 32000:
                v = v[:32000] + "  ...[truncated]"
            conv = as_date(v)
            if conv is None and v not in (None, "") and relation:
                typ = TYPES.get((relation, header[i - 1]), "")
                if typ.startswith(NUMERIC):
                    try:
                        v = int(v) if "." not in v else float(v)
                    except (TypeError, ValueError):
                        pass
            cell = ws.cell(row=r, column=i, value=conv[0] if conv else v)
            cell.font = MONO
            if conv:
                cell.number_format = conv[1]
    ws.freeze_panes = "A2"
    if header and rows:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(header))}{len(rows) + 1}"
    src = widths_from if widths_from is not None else ([header] + rows)
    for i in range(len(header)):
        longest = max((len(str(r[i])) for r in src if i < len(r) and r[i] is not None),
                      default=10)
        ws.column_dimensions[get_column_letter(i + 1)].width = min(max(longest + 2, 11), 58)
    ws.row_dimensions[1].height = 20


wb = Workbook()
wb.remove(wb.active)

# ---------------------------------------------------------------- _schema
cols_hdr, cols = csv_query("""
SELECT c.table_name, c.ordinal_position::int, c.column_name,
       format_type(a.atttypid, a.atttypmod) AS data_type,
       CASE WHEN c.is_nullable='NO' THEN 'NOT NULL' ELSE '' END AS nullable,
       coalesce(c.column_default,'') AS default_value
FROM information_schema.columns c
JOIN pg_class pc ON pc.relname = c.table_name AND pc.relnamespace='public'::regnamespace
JOIN pg_attribute a ON a.attrelid = pc.oid AND a.attname = c.column_name
WHERE c.table_schema='public'
ORDER BY c.table_name, c.ordinal_position""")

for _row in cols:
    TYPES[(_row[0], _row[2])] = _row[3]

key_hdr, keys = csv_query("""
SELECT cl.relname, a.attname,
       string_agg(DISTINCT CASE ct.contype WHEN 'p' THEN 'PK' WHEN 'u' THEN 'UQ'
                                           WHEN 'f' THEN 'FK' END, ' ')
FROM pg_constraint ct
JOIN pg_class cl ON cl.oid = ct.conrelid
JOIN unnest(ct.conkey) k(n) ON true
JOIN pg_attribute a ON a.attrelid = ct.conrelid AND a.attnum = k.n
WHERE ct.contype IN ('p','u','f') AND cl.relnamespace='public'::regnamespace
GROUP BY 1,2""")
KEYMAP = {(r[0], r[1]): r[2] for r in keys}

cnt_hdr, counts = csv_query(" UNION ALL ".join(
    f"SELECT '{t}' AS t, count(*)::text FROM {t}" for t, _ in TABLES))
COUNTS = {r[0]: r[1] for r in counts}

ws = wb.create_sheet("_schema")
hdr = ["table", "role", "rows", "#", "column", "type", "null?", "key", "default"]
data = []
for t, _ in TABLES:
    first = True
    for r in cols:
        if r[0] != t:
            continue
        data.append([t if first else "", ROLE.get(t, "") if first else "",
                     COUNTS.get(t, "") if first else "",
                     r[1], r[2], r[3], r[4], KEYMAP.get((t, r[2]), ""), r[5]])
        first = False
style(ws, hdr, data)
for row in ws.iter_rows(min_row=2, max_col=3):
    if row[0].value:
        for c in row:
            c.fill, c.font = SUB_FILL, Font(name="Consolas", size=10, bold=True)

# ---------------------------------------------------------------- _relationships
fk_hdr, fks = csv_query("""
SELECT conrelid::regclass::text || '.' || a.attname AS from_column,
       '-->' AS points_to,
       confrelid::regclass::text || '.' || af.attname AS to_column,
       CASE ct.confdeltype WHEN 'c' THEN 'ON DELETE CASCADE' ELSE '' END AS on_delete
FROM pg_constraint ct
JOIN unnest(ct.conkey)  WITH ORDINALITY k(n,i)  ON true
JOIN unnest(ct.confkey) WITH ORDINALITY fk(n,i) ON fk.i = k.i
JOIN pg_attribute a  ON a.attrelid  = ct.conrelid  AND a.attnum  = k.n
JOIN pg_attribute af ON af.attrelid = ct.confrelid AND af.attnum = fk.n
WHERE ct.contype='f'
ORDER BY 1""")
MEANING = {
    "email_tickets.email_id":            "which tickets this mail asked about",
    "email_tickets.ticket_number":       "the ticket this line item is for",
    "email_tickets.latest_reply_id":     "cache: which reply produced the latest status",
    "email_replies.email_id":            "which mail this reply answers (via In-Reply-To)",
    "reply_ticket_statuses.reply_id":    "which reply this remark arrived in",
    "reply_ticket_statuses.ticket_number": "the ticket this remark is about",
    "emails_sent.run_id":                "which pipeline run produced this mail",
    "tickets.first_run_id":              "which run first saw this ticket",
    "awb_registry.ticket_number":        "the keeper ticket for this AWB",
    "clickpost_statuses.run_id":         "which run fetched this courier status",
}
ws = wb.create_sheet("_relationships")
style(ws, ["from_column", "", "to_column", "on_delete", "what it means"],
      [[r[0], r[1], r[2], r[3], MEANING.get(r[0], "")] for r in fks])

# ---------------------------------------------------------------- one sheet per table
for t, order in TABLES:
    h, rows = csv_query(f"SELECT * FROM {t} ORDER BY {order}")
    ws = wb.create_sheet(t)
    style(ws, h, rows, relation=t)

wb.save(OUT)
print(f"wrote {OUT}")
for name in wb.sheetnames:
    n = wb[name].max_row - 1
    print(f"  {name:<24} {max(n, 0):>4} rows")
