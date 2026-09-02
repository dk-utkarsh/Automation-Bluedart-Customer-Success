"""Daily logistics pipeline: Zoho Desk tickets -> ClickPost status -> Mapping.xlsx

  python main.py                    # full run
  python main.py --today            # force a full-day ticket fetch (ignore watermark)
  python main.py --now              # cutoff = now instead of 17:30 IST
  python main.py --tickets <xlsx>   # reuse an existing export, skip the Desk fetch
  python main.py --skip-desk        # use the newest existing export
  python main.py --skip-clickpost   # part 1 only
  python main.py --asof 2026-08-19  # pin the date Delay Days is measured from
  python main.py --headed           # show the ClickPost browser (default: headless)
  python main.py --close            # quit Chrome when finished
  python main.py --selftest         # run the OTP-extraction checks, touch nothing
  python main.py --abandon-pending  # discard a stuck run so new tickets can flow
  python main.py --keep-files       # don't clean output/ after a successful run

  python main.py --watch            # stay live: reply -> ticket comments, 15:00 chases
  python main.py --process-replies  # one-shot: apply any replies waiting now
  python main.py --followups        # one-shot: send any due follow-ups

PART 1 - Zoho Desk
  Tickets created in the window are filtered to Logistics Team, still-open,
  Bluedart-only, then deduplicated on AWB -> output/Zoho_Desk_Logistics_<date>.xlsx

PART 2 - ClickPost
  Log in (OTP by email when the saved profile is not trusted), raise a bulk
  Shipment Lifecycle report for those AWBs, download it -> output/<id>_<date>.csv

PART 3 - Merge
  Join on AWB, drop Delivered / RTO / not-found, then emit output/Mapping.xlsx with
  AWB Number, Vinc Shipment EDD, Delay Days, Concern Type, Courier Partner, State.

PART 4 - Mail
  Mapping.xlsx goes to MAIL_TO. Sending is the commit point: the Desk watermark in
  state.json advances ONLY after the mail is accepted. A failure anywhere earlier
  leaves a `pending` block, and the next run re-processes that same export instead
  of fetching past it - so a broken evening cannot silently drop a day of tickets.

  Two cases finish WITHOUT mailing and still commit, because they are results rather
  than failures: no logistics tickets at all, and a Mapping file with zero rows
  (everything Delivered/RTO or unknown to ClickPost). Treating those as unfinished
  would leave a pending block nothing could ever satisfy.

  A run that keeps failing blocks newer tickets behind it by design. From the third
  attempt it says so loudly; --abandon-pending discards it deliberately.

  Once a run commits, its working files - the Desk export, the ClickPost CSV, the
  merged workbook, the Mapping file and any debug screenshots - are deleted, so each
  day starts from an empty output/. They are removed ONLY after the commit: while a
  retry is still possible the pending block points at the export it needs, and
  deleting early would turn a repeatable run into an unrecoverable one. A failed run
  therefore leaves everything on disk. --keep-files suppresses the cleanup.

Part 2 is skipped, not failed, when part 1 finds no AWBs: a day with no Bluedart
logistics tickets is a quiet day, not an error.
"""

import contextlib
import csv
import email
import imaplib
import io
import json
import os
import pathlib
import re
import shutil
import signal
import socket
import stat
import subprocess
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
import zipfile
from concurrent.futures import ThreadPoolExecutor
from datetime import date, datetime, timedelta, timezone
from email.header import decode_header, make_header
from email.utils import make_msgid, parsedate_to_datetime

from selenium import webdriver
from selenium.common.exceptions import TimeoutException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait

# ============================================================================
# Paths and shared configuration
# ============================================================================

BASE = pathlib.Path(__file__).parent

# db.py sits beside this file, but main.py is started by cron, by systemd, and
# imported by webhook_server.py - none of which reliably put its directory on
# sys.path. Anchor on __file__ rather than the working directory.
if str(BASE) not in sys.path:
    sys.path.insert(0, str(BASE))
import db                                                   # noqa: E402

ENV_FILE = BASE / ".env"
STATE_FILE = BASE / "state.json"
AWB_FILE = BASE / "awb_registry.json"
THREADS_FILE = BASE / "threads.json"
OUT = BASE / "output"
PROFILE_DIR = BASE / ".chrome-profile"
IS_WINDOWS = os.name == "nt"

# Chrome lives somewhere different on every platform, and on Linux the binary may be
# called google-chrome, google-chrome-stable or chromium. These are searched in order;
# CHROME_BINARY in .env overrides the search entirely. See chrome_path().
CHROME_CANDIDATES = [
    pathlib.Path(r"C:\Program Files\Google\Chrome\Application\chrome.exe"),
    pathlib.Path(r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe"),
] if IS_WINDOWS else [
    pathlib.Path("/usr/bin/google-chrome"),
    pathlib.Path("/usr/bin/google-chrome-stable"),
    pathlib.Path("/opt/google/chrome/chrome"),
    pathlib.Path("/usr/bin/chromium-browser"),
    pathlib.Path("/usr/bin/chromium"),
    pathlib.Path("/snap/bin/chromium"),
]
# chrome-for-testing publishes one build per platform, and the archive member is named
# differently on each - both are needed to fetch and unpack the matching driver.
DRIVER_PLATFORM = "win64" if IS_WINDOWS else "linux64"
DRIVER_NAME = "chromedriver.exe" if IS_WINDOWS else "chromedriver"
DRIVER_CACHE = (pathlib.Path.home() / ".cache" / "selenium" / "chromedriver"
                / DRIVER_PLATFORM)

IST = timezone(timedelta(hours=5, minutes=30))
PLACEHOLDER = "PASTE_VALUE_HERE"

ENV = {}            # filled by load_env() at start of main()


def load_env():
    env = {}
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line and not line.startswith("#") and "=" in line:
            k, v = line.split("=", 1)
            env[k.strip()] = v.strip()
    # Persistence is mirrored, not required: an unset DATABASE_URL leaves every
    # db.* call a no-op and the pipeline runs exactly as it did on JSON alone.
    db.configure(env.get("DATABASE_URL"))
    return env


def require_env(env, keys, hint):
    missing = [k for k in keys if not env.get(k) or env[k].strip() == PLACEHOLDER]
    if missing:
        print("These values are still unset in .env:")
        for k in missing:
            print("  - " + k)
        print(hint)
        raise SystemExit(1)


def norm_awb(v):
    """AWBs are typed by hand, so compare them without spaces or case."""
    return "".join(str(v or "").split()).upper()


def newest(pattern):
    files = sorted(OUT.glob(pattern), key=lambda p: p.stat().st_mtime)
    return files[-1] if files else None


def newest_export():
    """Newest raw Desk export. The *_with_status.xlsx files match the same glob but
    are pipeline OUTPUT - feeding one back in would merge a status column onto a
    sheet that already has one."""
    files = [p for p in OUT.glob("Zoho_Desk_Logistics_*.xlsx")
             if not p.stem.endswith("_with_status")]
    files.sort(key=lambda p: p.stat().st_mtime)
    return files[-1] if files else None


def banner(text):
    print("\n" + "=" * 68)
    print(text)
    print("=" * 68)


# ---------------------------------------------------------------------------
# Run state: the watermark is committed only once the mail is away
# ---------------------------------------------------------------------------
#
# state.json holds a COMMITTED watermark plus, while a run is in flight, a
# `pending` block. The watermark moves only after the mailer succeeds, so a run
# that dies in ClickPost or the merge leaves the window exactly where it was and
# the next run re-processes the same tickets instead of skipping them.
#
# A day whose result is genuinely empty commits too, without mailing: zero records
# is a finished run, not a failure. Without that, the first quiet day would leave a
# pending block nothing could ever satisfy and the pipeline would deadlock.


def load_state():
    if STATE_FILE.exists():
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    return {}


def save_state(state):
    STATE_FILE.write_text(json.dumps(state, indent=2), encoding="utf-8")


def start_pending(state, export_path, watermark, modified_utc=None):
    """Record the in-flight run. Deliberately does NOT touch last_created_utc."""
    prev = state.get("pending") or {}
    state["pending"] = {
        "export_file": str(export_path),
        "watermark_ticket": watermark["ticketNumber"],
        "watermark_utc": watermark["created_utc"].isoformat(),
        "started_ist": datetime.now(IST).isoformat(),
        "attempts": int(prev.get("attempts") or 0) + 1,
        "modified_utc": modified_utc,
        "stage_reached": "desk",
        "mailer_sent": False,
    }
    save_state(state)
    return state["pending"]


def mark_stage(state, stage):
    if state.get("pending"):
        state["pending"]["stage_reached"] = stage
        save_state(state)
    db.mark_stage(state.get("db_run_id"), stage)


def fail_current_run(state, error):
    """Record that the open run failed.

    Persistence only. The pending block in state.json is what actually drives
    the retry and is deliberately left exactly as it was - this just stops the
    run row sitting at 'pending' forever."""
    db.fail_run(state.get("db_run_id"), error)


def cleanup_run_files(paths, keep=False):
    """Delete this run's artefacts once the run is committed.

    Only ever called after the mail is away (or after a run that finished with
    nothing to send). While anything can still be retried the files must survive -
    the pending block points at the export the retry needs, so deleting early would
    make the run unrecoverable rather than merely repeatable.

    Debug screenshots go too, but only on success: on failure they are the evidence."""
    if keep:
        print("\n--keep-files: leaving {} file(s) in {}".format(len(paths), OUT.name))
        return
    removed = []
    for p in paths:
        if p is None:
            continue
        try:
            p = pathlib.Path(p)
            if p.exists():
                p.unlink()
                removed.append(p.name)
        except OSError as e:
            # A file open in Excel cannot be removed. Not fatal - the run is already
            # committed - but say so rather than pretending the folder is clean.
            print("  could not delete {}: {}".format(p.name, e.__class__.__name__))
    for png in OUT.glob("_*.png"):
        try:
            png.unlink()
            removed.append(png.name)
        except OSError:
            pass
    sup = OUT / "superseded"
    if sup.is_dir() and not any(sup.iterdir()):
        try:
            sup.rmdir()
        except OSError:
            pass
    if removed:
        print("\nCleaned up {} file(s): {}".format(len(removed), ", ".join(removed)))


def commit_run(state, ticket, created_utc, export_file, rows, mailed,
               modified_utc=None):
    """Advance the watermark and clear the pending block."""
    state["last_ticket_number"] = ticket
    state["last_created_utc"] = created_utc
    if modified_utc:
        # Advanced on the same terms as the created watermark: only once the mail is
        # away, so a failed run re-sweeps the same late edits rather than losing them.
        state["last_modified_utc"] = modified_utc
    state["last_run_ist"] = datetime.now(IST).isoformat()
    state["last_file"] = str(export_file) if export_file else None
    state["rows"] = rows
    state["last_mail_sent_ist"] = datetime.now(IST).isoformat() if mailed else \
        state.get("last_mail_sent_ist")
    state.pop("pending", None)
    # Mirrored before db_run_id is dropped, so the run row closes with the same
    # watermark the JSON just committed.
    db.commit_run(state.get("db_run_id"), ticket, created_utc,
                  modified_utc=modified_utc, rows_count=rows,
                  mailer_sent=mailed)
    state.pop("db_run_id", None)
    save_state(state)
    print("\nWatermark committed -> next run starts after #{}".format(ticket))


# ============================================================================
# PART 1 - Zoho Desk
# ============================================================================

# Desk API root. The host comes from ZOHO_API_DOMAIN in .env; the literal below is
# only the fallback for a .env written before that key existed. Resolved per call
# rather than at import, because ENV is not filled until main() runs.
DESK_API_DEFAULT = "https://desk.zoho.in"


def api_base():
    domain = (ENV.get("ZOHO_API_DOMAIN") or "").strip().rstrip("/")
    if not domain or domain == PLACEHOLDER:
        domain = DESK_API_DEFAULT
    return domain + "/api/v1/"


CUTOFF_HOUR, CUTOFF_MIN = 17, 30
DESK_ENV = ["ZOHO_CLIENT_ID", "ZOHO_CLIENT_SECRET", "ZOHO_REFRESH_TOKEN"]

# Excel columns: (header, key)
DESK_COLUMNS = [
    ("Ticket Number", "ticketNumber"),
    ("Created (IST)", "created"),
    ("Subject", "subject"),
    ("Status", "status"),
    ("Ticket Owner", "owner"),
    ("Department", "department"),
    ("Logistics Classification", "Logistics Classification"),
    ("AWB Number", "AWB Number"),
    ("Courier Partner", "Courier Partner"),
    ("Vinculum Shipment EDD", "Vinculum Shipment EDD"),
    ("Priority", "priority"),
    ("States", "States"),
    ("Pending with Department", "Pending with Department"),
]
# These live in the label-keyed customFields block. Their cf_* API names do NOT
# match their labels (e.g. "Vinculum Shipment EDD" is cf_logistics_follow_up_date),
# so always read them by display label.
PENDING_FIELD = "Pending with Department"
CF_FIELDS = ["Logistics Classification", "AWB Number", "Courier Partner",
             "Vinculum Shipment EDD", "States", PENDING_FIELD]

# The Desk picklist value is "Logistics Team", not "Logistics" - matching the bare
# word alone returns zero tickets. Both spellings are accepted so a picklist rename
# does not silently empty the report.
LOGISTICS_VALUES = {"logistics", "logistics team"}

# Zoho maps every status onto a statusType of Open / On Hold / Closed. Testing the
# type rather than the label means a renamed status cannot leak closed tickets back in.
CLOSED_TYPES = {"closed"}

# Courier names are entered inconsistently, so compare with spacing and case removed:
# "Blue Dart", "BlueDart" and "BLUEDART" all match.
COURIER_FIELD = "Courier Partner"
COURIER_VALUES = {"bluedart"}

TOKEN = None


def is_wanted_courier(row):
    """Report covers one courier only. A blank courier does not qualify."""
    return "".join(str(row.get(COURIER_FIELD) or "").split()).lower() in COURIER_VALUES


def is_open(row):
    """Exclude resolved tickets - the report is a live pending list."""
    return str(row.get("statusType") or "").strip().lower() not in CLOSED_TYPES


def has_logistics_data(row):
    """A ticket is a Logistics ticket when Pending with Department says so.

    This is deliberately the ONLY test. An earlier version treated any filled
    Logistics - Closure Info field as evidence, which pulled in tickets pending with
    other teams that merely happened to carry an AWB."""
    return str(row.get(PENDING_FIELD) or "").strip().lower() in LOGISTICS_VALUES


def load_registry():
    if AWB_FILE.exists():
        return json.loads(AWB_FILE.read_text(encoding="utf-8")).get("awbs", {})
    return {}


def save_registry(reg):
    AWB_FILE.write_text(json.dumps({"awbs": reg}, indent=2, sort_keys=True),
                        encoding="utf-8")


def dedupe_by_awb(rows, registry):
    """Keep the earliest-created ticket per AWB; report what was dropped.

    Tickets with no AWB are always kept - a blank is not a dedup key, so collapsing
    them would throw away unrelated tickets. Re-running the same window is safe: a
    registry hit on the SAME ticket number is the same ticket, not a duplicate."""
    kept, dropped, batch = [], [], {}
    for r in rows:                                    # already sorted oldest-first
        awb = norm_awb(r.get("AWB Number"))
        if not awb:
            kept.append(r)
            continue

        prev = registry.get(awb)
        if prev and str(prev["ticketNumber"]) != str(r["ticketNumber"]):
            if r["created_utc"] < datetime.fromisoformat(prev["created_utc"]):
                # Backfill turned up a genuinely earlier ticket - it becomes the keeper.
                registry[awb] = {"ticketNumber": r["ticketNumber"],
                                 "ticketId": r.get("ticketId"),
                                 "created_utc": r["created_utc"].isoformat()}
                kept.append(r)
            else:
                dropped.append({**r, "dupe_of": prev["ticketNumber"],
                                "dupe_scope": "earlier run"})
            continue

        if awb in batch:
            dropped.append({**r, "dupe_of": batch[awb]["ticketNumber"],
                            "dupe_scope": "this run"})
            continue

        batch[awb] = r
        kept.append(r)
        registry[awb] = {"ticketNumber": r["ticketNumber"],
                         "ticketId": r.get("ticketId"),
                         "created_utc": r["created_utc"].isoformat()}
    return kept, dropped


def get_token():
    body = urllib.parse.urlencode({
        "refresh_token": ENV["ZOHO_REFRESH_TOKEN"],
        "client_id": ENV["ZOHO_CLIENT_ID"],
        "client_secret": ENV["ZOHO_CLIENT_SECRET"],
        "grant_type": "refresh_token",
    }).encode()
    url = ENV.get("ZOHO_ACCOUNTS_URL", "https://accounts.zoho.in") + "/oauth/v2/token"
    with urllib.request.urlopen(urllib.request.Request(url, data=body), timeout=30) as r:
        data = json.load(r)
    if "access_token" not in data:
        raise SystemExit("Token refresh failed: {}".format(data))
    return data["access_token"]


def api_get(path, retries=4):
    """GET with token refresh on 401 and backoff on 429/5xx."""
    global TOKEN
    for attempt in range(retries):
        headers = {"Authorization": "Zoho-oauthtoken " + TOKEN}
        org = (ENV.get("ZOHO_ORG_ID") or "").strip()
        if org and org != PLACEHOLDER:
            headers["orgId"] = org        # optional: single-org tokens resolve it
        req = urllib.request.Request(api_base() + path, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=45) as r:
                return json.load(r)
        except urllib.error.HTTPError as e:
            if e.code == 204:
                return {}
            if e.code == 401 and attempt < retries - 1:
                TOKEN = get_token()
                continue
            if e.code in (429, 500, 502, 503) and attempt < retries - 1:
                time.sleep(2 ** attempt)
                continue
            raise
        except (urllib.error.URLError, TimeoutError):
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
                continue
            raise
    raise RuntimeError("unreachable")


def parse_zoho_time(s):
    """Zoho returns UTC as 2026-08-18T11:13:49.000Z"""
    return datetime.strptime(s, "%Y-%m-%dT%H:%M:%S.%fZ").replace(tzinfo=timezone.utc)


def resolve_window(force_today, force_now):
    now_ist = datetime.now(IST)
    cutoff_ist = now_ist.replace(hour=CUTOFF_HOUR, minute=CUTOFF_MIN,
                                 second=0, microsecond=0)
    if force_now or now_ist < cutoff_ist:
        cutoff_ist = now_ist

    state = {}
    if STATE_FILE.exists():
        state = json.loads(STATE_FILE.read_text(encoding="utf-8"))

    if not force_today and state.get("last_created_utc"):
        # Incremental: strictly after the last ticket already exported.
        start_utc = datetime.fromisoformat(state["last_created_utc"])
        mode = "incremental (after #{} @ {:%Y-%m-%d %H:%M} IST)".format(
            state.get("last_ticket_number"), start_utc.astimezone(IST))
        inclusive = False
    else:
        start_utc = now_ist.replace(hour=0, minute=0, second=0,
                                    microsecond=0).astimezone(timezone.utc)
        mode = "full day (today 00:00 IST)"
        inclusive = True
    return start_utc, cutoff_ist.astimezone(timezone.utc), cutoff_ist, mode, inclusive


def collect_tickets(start_utc, end_utc, inclusive):
    """Page newest-first until we fall out of the window."""
    found, frm = [], 1
    while True:
        data = api_get("tickets?limit=100&from={}&sortBy=-createdTime".format(frm))
        rows = data.get("data") or []
        if not rows:
            break
        done = False
        for t in rows:
            ct = parse_zoho_time(t["createdTime"])
            if ct > end_utc:
                continue                      # created after the cutoff - skip
            if ct < start_utc or (not inclusive and ct <= start_utc):
                done = True
                break
            found.append(t)
        if done or len(rows) < 100:
            break
        frm += 100
    found.sort(key=lambda t: t["createdTime"])
    return found


# The list endpoint does NOT return modifiedTime unless it is named in `fields`.
# sortBy=-modifiedTime alone sorts correctly but hands back rows without the field,
# so the stop condition below would read None on every row and never terminate
# properly. Verified against the live API on 2026-08-24.
MODIFIED_FIELDS = "id,ticketNumber,createdTime,modifiedTime,statusType"

# How far back the late-edit sweep will look, by CREATION date. Measured on live data
# 2026-08-24: 1117 tickets were created earlier and touched that day, but only 179 of
# them were under a fortnight old - a single bulk update had brushed ~890 tickets aged
# 14-30 days. Without this bound every run would fetch detail for all 1117.
#
# The trade-off is explicit: a ticket rerouted to Logistics MORE than this many days
# after it was raised is still missed. Fourteen days is well past the point a shipment
# escalation is still live. Override with LATE_EDIT_LOOKBACK_DAYS in .env.
LATE_EDIT_LOOKBACK_DAYS = 14


def resolve_modified_start(state, created_start_utc):
    """Where the modified-since sweep begins.

    Falls back to the created watermark, so the very first sweep looks back exactly
    as far as the created one does rather than over the whole history."""
    v = state.get("last_modified_utc")
    return datetime.fromisoformat(v) if v else created_start_utc


def collect_modified(start_utc, end_utc):
    """Tickets whose modifiedTime falls in (start_utc, end_utc], newest first.

    Why this exists: the created-time watermark only ever moves forward, so a ticket
    created before it and rerouted to Logistics afterwards would never be looked at
    again - it would be silently absent from every future report. Sweeping by
    modifiedTime is the only way to see those."""
    found, frm = [], 1
    while True:
        data = api_get("tickets?limit=100&from={}&sortBy=-modifiedTime&fields={}".format(
            frm, MODIFIED_FIELDS))
        rows = data.get("data") or []
        if not rows:
            break
        done = False
        for t in rows:
            raw = t.get("modifiedTime")
            if not raw:
                continue                      # no timestamp to judge it by
            mt = parse_zoho_time(raw)
            if mt > end_utc:
                continue                      # modified after the cutoff - skip
            if mt <= start_utc:
                done = True
                break
            found.append(t)
        if done or len(rows) < 100:
            break
        frm += 100
    return found


def reported_tickets():
    """Ticket numbers that have already gone out in a mail, read from threads.json.

    The AWB registry cannot answer this. A registry hit on the SAME ticket number is
    deliberately NOT treated as a duplicate - that is what makes re-running a day
    reproduce its rows instead of emptying them. So without this set, a ticket that
    has already been escalated would be mailed again every time somebody touched it,
    which is precisely what the modified sweep would otherwise cause."""
    out = set()
    for t in load_threads():
        out.update(str(n) for n in thread_targets(t))
    return out


def fetch_detail(t):
    d = api_get("tickets/{}?include=assignee,departments".format(t["id"]))
    cf = d.get("customFields") or {}
    a = d.get("assignee") or {}
    owner = " ".join(x for x in [a.get("firstName"), a.get("lastName")] if x).strip() \
        or a.get("email", "")
    row = {
        # The internal id is what every write endpoint needs, and it cannot be derived
        # from the ticket number later - the search API is outside our OAuth scope.
        # Captured here so it can ride along into the AWB registry.
        "ticketId": d.get("id"),
        "ticketNumber": d.get("ticketNumber"),
        "created": parse_zoho_time(d["createdTime"]).astimezone(IST)
                                                    .strftime("%Y-%m-%d %H:%M:%S"),
        "created_utc": parse_zoho_time(d["createdTime"]),
        "subject": d.get("subject") or "",
        "status": d.get("status") or "",
        "statusType": d.get("statusType") or "",
        "owner": owner or "Unassigned",
        "department": (d.get("department") or {}).get("name", ""),
        "priority": d.get("priority") or "",     # standard field, not a custom field
    }
    for f in CF_FIELDS:
        v = cf.get(f)
        row[f] = "" if v is None else str(v)
    return row


def write_desk_excel(rows, dropped, cutoff_ist, path, n_closed=0, n_courier=0):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Logistics Closure Info"

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E79")
    ws.append([h for h, _ in DESK_COLUMNS])
    for c in ws[1]:
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    for r in rows:
        ws.append([r.get(k, "") for _, k in DESK_COLUMNS])

    widths = [15, 20, 46, 14, 22, 30, 26, 18, 18, 22, 12, 16, 24]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(
        get_column_letter(len(DESK_COLUMNS)), ws.max_row)
    for row in ws.iter_rows(min_row=2):
        row[2].alignment = Alignment(vertical="top", wrap_text=False)

    dup = wb.create_sheet("Duplicates Dropped")
    dup.append(["Ticket Number", "Created (IST)", "AWB Number",
                "Duplicate Of", "Detected In", "Subject"])
    for c in dup[1]:
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r in dropped:
        dup.append([r.get("ticketNumber"), r.get("created"), r.get("AWB Number"),
                    r.get("dupe_of"), r.get("dupe_scope"), r.get("subject")])
    for i, w in enumerate([15, 20, 18, 15, 14, 46], 1):
        dup.column_dimensions[get_column_letter(i)].width = w
    dup.freeze_panes = "A2"

    info = wb.create_sheet("Run Info")
    info.append(["Generated (IST)", datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")])
    info.append(["Window cutoff (IST)", cutoff_ist.strftime("%Y-%m-%d %H:%M")])
    info.append(["Tickets exported", len(rows)])
    info.append(["Duplicates dropped", len(dropped)])
    info.append(["Closed/solved excluded", n_closed])
    info.append(["Non-Bluedart excluded", n_courier])
    if rows:
        info.append(["First ticket", rows[0]["ticketNumber"]])
        info.append(["Last ticket", rows[-1]["ticketNumber"]])
    info.column_dimensions["A"].width = 24
    info.column_dimensions["B"].width = 24
    for c in info["A"]:
        c.font = Font(bold=True)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def run_desk_fetch(state, force_today=False, force_now=False):
    """Fetch and export logistics tickets.

    Returns {"path", "watermark"} for a run that still has work to do, or None when
    there is nothing to process. The watermark is NOT committed here - that happens
    only after the mail is away, so a later failure re-processes these same tickets."""
    global TOKEN
    require_env(ENV, DESK_ENV,
                "Client id/secret come from api-console.zoho.in; "
                "org id: Desk > Setup > Organization details.")

    start_utc, end_utc, cutoff_ist, mode, inclusive = resolve_window(force_today, force_now)
    print("Mode   : {}".format(mode))
    print("Window : {:%Y-%m-%d %H:%M} -> {:%Y-%m-%d %H:%M} IST".format(
        start_utc.astimezone(IST), cutoff_ist))

    # Opened here rather than at the export, so a run that finds nothing - or
    # dies in the fetch - is still on the record. The id rides in state.json so
    # a resumed run closes the row it opened instead of starting a second one.
    run_db_id = db.begin_run(watermark_ticket=state.get("last_ticket_number"),
                             watermark_utc=state.get("last_created_utc"))
    if run_db_id:
        state["db_run_id"] = run_db_id
        save_state(state)

    TOKEN = get_token()
    tickets = collect_tickets(start_utc, end_utc, inclusive)
    print("Tickets in window: {}".format(len(tickets)))

    # Second sweep, by modifiedTime, for tickets created BEFORE this window but
    # touched inside it - the ones a creation-time watermark can never revisit.
    # Three exclusions keep it to exactly those: anything the first sweep already
    # returned, anything created inside the window, and anything already mailed.
    mod_start = resolve_modified_start(state, start_utc)
    swept = collect_modified(mod_start, end_utc)
    modified_mark = max((parse_zoho_time(t["modifiedTime"]) for t in swept),
                        default=None)
    fresh_ids = {t["id"] for t in tickets}
    already = reported_tickets()
    try:
        lookback = int(ENV.get("LATE_EDIT_LOOKBACK_DAYS") or LATE_EDIT_LOOKBACK_DAYS)
    except ValueError:
        lookback = LATE_EDIT_LOOKBACK_DAYS
    floor_utc = end_utc - timedelta(days=lookback)
    late, too_old, closed = [], 0, 0
    for t in swept:
        if t["id"] in fresh_ids:
            continue                       # the first sweep already has it
        ct = parse_zoho_time(t["createdTime"])
        if ct > start_utc:
            continue                       # created inside the window, already covered
        if (t.get("statusType") or "") == "Closed":
            closed += 1                    # only skip on an explicit Closed
            continue
        if ct < floor_utc:
            too_old += 1
            continue
        if str(t.get("ticketNumber") or "") in already:
            continue
        late.append(t)
    print("Modified since {:%Y-%m-%d %H:%M} IST: {}  -> {} late edit(s) to examine"
          .format(mod_start.astimezone(IST), len(swept), len(late)))
    # Printed every run so the bound stays visible rather than silently dropping work.
    if too_old or closed:
        print("  skipped: {} older than the {}-day lookback, {} closed".format(
            too_old, lookback, closed))
    if late:
        print("  late edits: {}".format(
            ", ".join("#" + str(t.get("ticketNumber")) for t in late[:12])))

    if not tickets and not late:
        print("Nothing new - no file written.")
        return None

    print("Fetching custom fields...")
    with ThreadPoolExecutor(max_workers=5) as pool:
        rows = list(pool.map(fetch_detail, tickets + late))
    rows.sort(key=lambda r: r["created_utc"])

    # Every ticket this run EXAMINED, before any filter. "Processed" means looked
    # at, not merely mailed: recording only the survivors left no way to answer
    # why a ticket never reached a report.
    for r in rows:
        db.upsert_ticket(r["ticketNumber"], ticket_id=r.get("ticketId"),
                         awb=norm_awb(r.get("AWB Number")) or None,
                         created_utc=r["created_utc"],
                         extra={"courier": r.get(COURIER_FIELD),
                                "logistics_class": r.get("Logistics Classification"),
                                "states": r.get("States")},
                         run_id=run_db_id)

    # The created watermark is the last ticket SEEN by the FIRST sweep, before any
    # filtering, so that non-logistics tickets at its tail are not re-scanned every
    # run. Late-edit tickets must never set it: they are older than the window, so
    # letting them would drag the watermark backwards and re-scan days of history.
    created_rows = [r for r in rows if r["ticketId"] in fresh_ids]
    if created_rows:
        watermark = created_rows[-1]
    else:
        # Only late edits this time - hold the created watermark exactly where it was.
        watermark = {"ticketNumber": state.get("last_ticket_number"),
                     "created_utc": datetime.fromisoformat(state["last_created_utc"])}

    logistics = [r for r in rows if has_logistics_data(r)]
    print("Logistics tickets: {} of {} ({} = Logistics Team)".format(
        len(logistics), len(rows), PENDING_FIELD))
    others = sorted({str(r.get(PENDING_FIELD) or "").strip() for r in rows
                     if str(r.get(PENDING_FIELD) or "").strip()
                     and not has_logistics_data(r)})
    if others:
        print("  other departments seen (excluded): {}".format(", ".join(others)))

    # Status filter runs BEFORE dedupe on purpose: if the earliest ticket for an AWB
    # is already resolved, the live pending duplicate is the one worth surfacing.
    n_before = len(logistics)
    logistics = [r for r in logistics if is_open(r)]
    n_closed = n_before - len(logistics)
    print("After status filter: {}  ({} closed/solved excluded)".format(
        len(logistics), n_closed))

    n_before = len(logistics)
    logistics = [r for r in logistics if is_wanted_courier(r)]
    n_courier = n_before - len(logistics)
    print("After courier filter: {}  ({} non-Bluedart excluded)".format(
        len(logistics), n_courier))

    registry = load_registry()
    kept, dropped = dedupe_by_awb(logistics, registry)
    print("After AWB dedupe : {}  ({} duplicate(s) dropped)".format(
        len(kept), len(dropped)))

    if not kept and not dropped:
        # Nothing to report and nothing to mail - a finished run, so commit.
        commit_run(state, watermark["ticketNumber"],
                   watermark["created_utc"].isoformat(), None, 0, mailed=False,
                   modified_utc=modified_mark.isoformat() if modified_mark else None)
        print("No logistics tickets in window - watermark advanced, no file written.")
        return None

    # Exactly one file per day. Re-running replaces today's report rather than piling
    # up _run2, _run3... - but the replaced version moves into output/superseded/
    # rather than being deleted, so a bad run is always recoverable.
    stamp = cutoff_ist.strftime("%Y-%m-%d")
    out = OUT / "Zoho_Desk_Logistics_{}.xlsx".format(stamp)
    if out.exists():
        archive = OUT / "superseded"
        archive.mkdir(parents=True, exist_ok=True)
        prev = archive / "Zoho_Desk_Logistics_{}_replaced_{:%H%M%S}.xlsx".format(
            stamp, datetime.now(IST))
        try:
            out.replace(prev)
            print("Replaced earlier version -> superseded/{}".format(prev.name))
        except PermissionError:
            # Windows locks an open workbook. A scheduled run must not lose a day of
            # data because someone left the file open in Excel, so fall back to a
            # timestamped name and say so loudly.
            out = OUT / "Zoho_Desk_Logistics_{}_{:%H%M%S}.xlsx".format(
                stamp, datetime.now(IST))
            print("WARNING: {} report is open in another program (Excel?).".format(stamp))
            print("         Could not replace it - writing {} instead.".format(out.name))

    write_desk_excel(kept, dropped, cutoff_ist, out, n_closed, n_courier)
    save_registry(registry)

    # Held, not committed: the watermark only moves once the mail is away.
    pending = start_pending(state, out, watermark,
                            modified_mark.isoformat() if modified_mark else None)

    print("\nSaved: {}".format(out))
    print("Rows : {}  (#{} -> #{})".format(
        len(kept), kept[0]["ticketNumber"], kept[-1]["ticketNumber"]))
    print("State: watermark #{} held pending the mailer (attempt {})".format(
        watermark["ticketNumber"], pending["attempts"]))
    print("AWBs : {} tracked in {}".format(len(registry), AWB_FILE.name))
    return {"path": out, "watermark": watermark,
            "modified_utc": modified_mark.isoformat() if modified_mark else None}


# ============================================================================
# PART 2a - OTP retrieval over IMAP
# ============================================================================
#
# Where the mail lands: ClickPost OTPs are filed into the "Notification" folder,
# NOT INBOX (a Zoho filter moves them). Searching INBOX alone finds nothing and the
# login times out, so both are searched; override with OTP_FOLDER in .env.
#
# What the code looks like: the OTP is ALPHANUMERIC, not numeric - observed values
# W8BV3V, X63YUN, YHUM5T:
#     Your One-Time Password (OTP) for ClickPost login is: X63YUN
# A digits-only pattern silently never matches. A mail is only mined for a code once
# it reads like an OTP mail (OTP_CONTEXT_RE) - without that gate the ClickPost
# *report* mails, which carry "Report ID: 361870", hand back a report id as the OTP.
#
# Which mail: not the newest by timestamp. Date headers come from Zoho's servers
# while the login runs on this PC, so clock skew makes a "newer than t0" test either
# miss the real OTP or accept a stale one. Instead the UIDs present BEFORE the
# password is submitted are snapshotted, and a UID absent from that snapshot wins.

SENDER_MATCH = "clickpost"
DEFAULT_FOLDERS = ["Notification", "INBOX"]
IMAP_TIMEOUT = 15
IMAP_TRIES = 8
IMAP_BACKOFF = 2
MAIL_ENV = ["OTP_EMAIL", "OTP_EMAIL_PASSWORD", "IMAP_HOST", "IMAP_PORT"]

OTP_CONTEXT_RE = re.compile(
    r"one[\s-]*time\s*password|\botp\b|verification\s*code|security\s*code|passcode",
    re.IGNORECASE)
# Anchored forms: each pins the code to a label that precedes it, so the match is
# positional and needs no digit. The scoped (?i:...) keeps the LABEL
# case-insensitive while the captured code stays strictly [A-Z0-9] - a plain
# re.IGNORECASE would relax the capture too and start matching lowercase prose.
ANCHORED_RES = [
    re.compile(r"\bis\s*:?\s+([A-Z0-9]{4,8})\b"),
    re.compile(r"\b(?i:otp|code|passcode|password)\s*[:\-]\s*([A-Z0-9]{4,8})\b"),
]
CANDIDATE_RE = re.compile(r"(?<![A-Za-z0-9])([A-Z0-9]{4,8})(?![A-Za-z0-9])")
# Words that legitimately appear in capitals inside OTP mails. Without this the
# loose scan could return one of them as the "code" now that letters-only codes
# are accepted.
CODE_STOPWORDS = {
    "OTP", "HTTP", "HTTPS", "HTML", "EMAIL", "LOGIN", "USER", "PASS", "CODE",
    "PASSWORD", "PASSCODE", "VALID", "ONLY", "THIS", "YOUR", "PLEASE", "NOTE",
    "TEAM", "INDIA", "TIME", "DEAR", "HELLO", "THANKS", "REPLY", "SHARE",
    "NEVER", "MINUTES", "SECURITY", "VERIFY", "ACCOUNT", "SUPPORT", "NOREPLY",
}
YEAR_RE = re.compile(r"^(19|20)\d{2}$")


def html_to_text(s):
    """Flatten an HTML part enough to read the code out of it."""
    s = re.sub(r"(?is)<(script|style)[^>]*>.*?</\1>", " ", s)
    s = re.sub(r"<[^>]+>", " ", s)
    s = (s.replace("&nbsp;", " ").replace("&amp;", "&")
          .replace("&lt;", "<").replace("&gt;", ">").replace("&#39;", "'"))
    return re.sub(r"\s+", " ", s).strip()


def _plausible_code(tok, anchored=False):
    """Could this token be the code?

    A code may be ANY mix of letters and digits, letters-only included - ClickPost
    issues those (KWUUHG, seen 2026-08-24). So a digit is never required. What is
    rejected is only what cannot be a code: a bare year, and the handful of English
    words that appear in capitals in these mails.

    `anchored` says the token was pinned to a label ("... OTP is: X"), which makes
    it trustworthy enough to skip the stopword check - a code legitimately spelling
    one of those words should still win when the mail points straight at it."""
    if YEAR_RE.match(tok):
        return False
    if anchored:
        return True
    return tok.upper() not in CODE_STOPWORDS


def extract_otp(text):
    """Pull the OTP out of a mail body. Returns the code, or None."""
    if not text:
        return None
    flat = re.sub(r"\s+", " ", text)

    # Gate: report mails and newsletters must never yield a "code".
    if not OTP_CONTEXT_RE.search(flat):
        return None

    for rx in ANCHORED_RES:
        m = rx.search(flat)
        if m and _plausible_code(m.group(1), anchored=True):
            return m.group(1)

    # Nothing was labelled, so fall back to scanning. Both shapes are accepted, but
    # one carrying a digit is far less likely to be an ordinary word, so it wins if
    # present; a letters-only token is still returned when that is all there is.
    cands = [t for t in CANDIDATE_RE.findall(flat) if _plausible_code(t)]
    for tok in cands:
        if any(c.isdigit() for c in tok):
            return tok
    return cands[0] if cands else None


def message_text(msg):
    """Subject plus text/plain, falling back to flattened HTML."""
    plain, html = [], []
    for part in msg.walk():
        ctype = part.get_content_type()
        if ctype not in ("text/plain", "text/html"):
            continue
        try:
            body = part.get_payload(decode=True)
        except Exception:
            continue
        if not body:
            continue
        body = body.decode(part.get_content_charset() or "utf-8", "ignore")
        (plain if ctype == "text/plain" else html).append(body)

    subject = str(make_header(decode_header(msg.get("Subject") or "")))
    parts = [subject] + plain + [html_to_text(h) for h in html]
    return "\n".join(p for p in parts if p)


def mail_folders():
    configured = (ENV.get("OTP_FOLDER") or "").strip()
    if configured:
        return [f.strip() for f in configured.split(",") if f.strip()]
    return list(DEFAULT_FOLDERS)


def imap_connect(tries=IMAP_TRIES, quiet=False):
    """IMAP_SSL connect + login, retrying dropped handshakes.

    imap.zoho.in on this network intermittently times out on connect - roughly one
    attempt in three - so a single failure must not end the login."""
    last = None
    for attempt in range(1, tries + 1):
        try:
            M = imaplib.IMAP4_SSL(ENV["IMAP_HOST"], int(ENV["IMAP_PORT"]),
                                  timeout=IMAP_TIMEOUT)
            M.login(ENV["OTP_EMAIL"], ENV["OTP_EMAIL_PASSWORD"])
            return M
        except Exception as e:
            last = e
            if not quiet:
                print("    IMAP connect {}/{}: {}".format(
                    attempt, tries, e.__class__.__name__))
            time.sleep(IMAP_BACKOFF)
    raise RuntimeError("IMAP unreachable after {} attempts: {}".format(tries, last))


def _uids_in(M, folder):
    typ, _ = M.select('"{}"'.format(folder), readonly=True)
    if typ != "OK":
        return set()
    typ, data = M.uid("SEARCH", None, '(FROM "{}")'.format(SENDER_MATCH))
    if typ != "OK":
        return set()
    return set((data[0] or b"").split())


def mail_snapshot():
    """ClickPost UIDs per folder, taken BEFORE the password is submitted."""
    M = imap_connect()
    try:
        base = {f: _uids_in(M, f) for f in mail_folders()}
        print("    baseline: " + ", ".join(
            "{}={}".format(f, len(u)) for f, u in base.items()))
        return base
    finally:
        try:
            M.logout()
        except Exception:
            pass


def wait_for_otp(baseline, timeout=300, poll=5):
    """Block until a ClickPost mail absent from `baseline` yields an OTP.

    The connection is held open across polls and re-established on failure;
    reconnecting from scratch every poll spends the whole budget on handshakes."""
    deadline = time.time() + timeout
    announced = False
    M = None

    while time.time() < deadline:
        try:
            if M is None:
                M = imap_connect(tries=3, quiet=True)
            for folder in mail_folders():
                fresh = _uids_in(M, folder) - baseline.get(folder, set())
                # Newest first: every login attempt mints a fresh OTP and invalidates
                # the previous one, so the highest UID is the only usable code.
                for uid in sorted(fresh, key=lambda u: int(u), reverse=True):
                    typ, d = M.uid("FETCH", uid, "(RFC822)")
                    if not d or not d[0]:
                        continue
                    msg = email.message_from_bytes(d[0][1])
                    subject = str(make_header(decode_header(msg.get("Subject") or "")))
                    print("    new mail [{}]: {} | {}".format(
                        folder, msg.get("From"), subject))
                    code = extract_otp(message_text(msg))
                    if code:
                        return code
                    print("      not an OTP mail - still waiting")
                    baseline.setdefault(folder, set()).add(uid)
        except Exception as e:
            print("    (mailbox hiccup: {} - reconnecting)".format(e.__class__.__name__))
            try:
                M.logout()
            except Exception:
                pass
            M = None

        if not announced:
            print("    waiting for the OTP mail (up to {}s)...".format(timeout))
            announced = True
        time.sleep(poll)

    if M is not None:
        try:
            M.logout()
        except Exception:
            pass
    raise TimeoutError("No ClickPost OTP mail arrived within {}s".format(timeout))


# ============================================================================
# PART 2b - Chrome / ClickPost login
# ============================================================================

USERNAME_SEL = (By.CSS_SELECTOR, "#input-username")
PASSWORD_SEL = (By.CSS_SELECTOR, "#input-password")
OTP_TIMEOUT = 300
# Chrome runs with a debugging port so a later run can attach to the SAME logged-in
# window instead of logging in again and burning another OTP.
DEBUG_PORT = 9222
CLICKPOST_ENV = ["CLICKPOST_URL", "CLICKPOST_EMAIL", "CLICKPOST_PASSWORD"]


def chrome_path():
    """The Chrome binary to drive.

    CHROME_BINARY in .env wins, so an unusual install needs no code change. Otherwise
    the known locations for this platform are tried, then PATH."""
    override = (ENV.get("CHROME_BINARY") or "").strip()
    if override:
        exe = pathlib.Path(override)
        if not exe.exists():
            raise SystemExit(
                "CHROME_BINARY is set to {} but there is no such file".format(exe))
        return exe
    for cand in CHROME_CANDIDATES:
        if cand.exists():
            return cand
    for name in ("google-chrome", "google-chrome-stable", "chromium",
                 "chromium-browser"):
        found = shutil.which(name)
        if found:
            return pathlib.Path(found)
    raise SystemExit(
        "Could not find Chrome. Install google-chrome-stable, or point CHROME_BINARY "
        "in .env at the binary.")


def chrome_version():
    """Product version of the Chrome that will actually be driven, e.g. 151.0.7922.170.

    Windows has no --version flag that prints to stdout reliably, hence the PowerShell
    property read; Linux does, so it is asked directly and the number pulled out of
    "Google Chrome 151.0.7922.170"."""
    exe = chrome_path()
    if IS_WINDOWS:
        out = subprocess.run(
            ["powershell", "-NoProfile", "-Command",
             "(Get-Item '{}').VersionInfo.ProductVersion".format(exe)],
            capture_output=True, text=True)
        v = out.stdout.strip()
    else:
        out = subprocess.run([str(exe), "--version"],
                             capture_output=True, text=True)
        found = re.search(r"(\d+\.\d+\.\d+\.\d+)", out.stdout or "")
        v = found.group(1) if found else ""
    if not re.match(r"^\d+\.\d+\.\d+\.\d+$", v):
        raise SystemExit(
            "Could not read the Chrome version from {} (got {!r})".format(exe, v))
    return v


def ensure_chromedriver(tries=5):
    """Return a chromedriver matching the installed Chrome, downloading if needed.

    Selenium Manager cannot be relied on here - its downloader intermittently fails on
    this network and silently falls back to a CACHED, VERSION-MISMATCHED driver, which
    dies at launch with a confusing "only supports Chrome version N" error."""
    ver = chrome_version()
    exe = DRIVER_CACHE / ver / DRIVER_NAME
    if exe.exists():
        return exe

    url = ("https://storage.googleapis.com/chrome-for-testing-public/"
           "{ver}/{plat}/chromedriver-{plat}.zip".format(
               ver=ver, plat=DRIVER_PLATFORM))
    print("  chromedriver {} not cached - downloading".format(ver))
    blob = None
    for attempt in range(1, tries + 1):
        try:
            blob = urllib.request.urlopen(url, timeout=120).read()
            break
        except Exception as e:
            print("    attempt {}/{} failed: {}".format(
                attempt, tries, e.__class__.__name__))
            time.sleep(3)
    if blob is None:
        raise SystemExit("Could not download chromedriver {} from {}".format(ver, url))

    exe.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(io.BytesIO(blob)) as z:
        for member in z.namelist():
            # Exact basename, not endswith: the archive also carries
            # LICENSE.chromedriver, which an endswith("chromedriver") test would
            # happily extract as the driver on Linux.
            if member.rsplit("/", 1)[-1] == DRIVER_NAME:
                exe.write_bytes(z.read(member))
                break
    if not exe.exists():
        raise SystemExit(
            "{} was not found inside {}".format(DRIVER_NAME, url))
    if not IS_WINDOWS:
        # zipfile does not preserve the executable bit, so without this the driver
        # fails to launch with a bare PermissionError that names no cause.
        exe.chmod(exe.stat().st_mode | stat.S_IXUSR | stat.S_IXGRP | stat.S_IXOTH)
    print("  installed {}".format(exe))
    return exe


def close_stale_chrome():
    """A previous detached run still holds .chrome-profile, and Chrome refuses to
    start a second instance on the same profile. Only processes launched against THIS
    profile directory are closed - ordinary browsing windows are untouched.

    This matters more on a server than on a desktop: runs are detached on purpose so
    the ClickPost session survives between them, so without a reaper a crashed run
    would leave the profile locked and block every run after it."""
    killed = []
    if IS_WINDOWS:
        script = (
            "Get-CimInstance Win32_Process -Filter \"Name='chrome.exe'\" | "
            "Where-Object { $_.CommandLine -like '*" + str(PROFILE_DIR) + "*' } | "
            "ForEach-Object { Stop-Process -Id $_.ProcessId -Force; $_.ProcessId }")
        out = subprocess.run(["powershell", "-NoProfile", "-Command", script],
                             capture_output=True, text=True)
        killed = [x for x in out.stdout.split() if x.strip().isdigit()]
    else:
        # pgrep -f matches against the whole command line, which is the same
        # "only processes holding OUR profile" test as the Windows branch.
        try:
            out = subprocess.run(["pgrep", "-f", str(PROFILE_DIR)],
                                 capture_output=True, text=True)
        except FileNotFoundError:
            print("  note: pgrep not installed - cannot reap stale Chrome processes")
            return
        mine = os.getpid()
        for tok in out.stdout.split():
            if not tok.strip().isdigit() or int(tok) == mine:
                continue
            try:
                os.kill(int(tok), signal.SIGKILL)
                killed.append(tok)
            except (ProcessLookupError, PermissionError):
                pass
    if killed:
        print("  closed {} stale Chrome process(es) holding the profile".format(
            len(killed)))
        time.sleep(2)


def start_browser(keep_open=True, headless=None):
    """Launch Chrome on the saved profile.

    Headless by default so a scheduled run doesn't throw a window onto someone's
    desktop; --headed brings it back for debugging. A fixed window size is set
    explicitly because --start-maximized does nothing without a desktop, and the
    ClickPost report builder needs a wide viewport for its Fields panel."""
    if headless is None:
        headless = not flag("--headed")
    close_stale_chrome()
    opts = Options()
    opts.add_argument("--user-data-dir={}".format(PROFILE_DIR))
    opts.add_argument("--remote-debugging-port={}".format(DEBUG_PORT))
    if headless:
        opts.add_argument("--headless=new")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--disable-gpu")
    else:
        opts.add_argument("--start-maximized")
    if not IS_WINDOWS:
        # Both are required on a server. The sandbox needs privileges a service
        # account does not have, and the default 64MB /dev/shm makes Chrome die
        # part-way through a long report run.
        opts.add_argument("--no-sandbox")
        opts.add_argument("--disable-dev-shm-usage")
    # Pinned explicitly so the driver launches the same binary chrome_version() was
    # measured against - otherwise a second Chrome on PATH can cause a version clash.
    opts.binary_location = str(chrome_path())
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    # Chrome offered to save these credentials on the first successful run, then
    # autofilled them on the next one - and the typed value landed on top of the
    # autofilled one, producing a doubled password string and an auth failure.
    opts.add_experimental_option("prefs", {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.password_manager_leak_detection": False,
        "autofill.profile_enabled": False,
    })
    if keep_open:
        opts.add_experimental_option("detach", True)
    return webdriver.Chrome(service=Service(str(ensure_chromedriver())), options=opts)


def attach():
    """Attach to a Chrome already running on DEBUG_PORT, or None if there isn't one."""
    try:
        with socket.create_connection(("127.0.0.1", DEBUG_PORT), timeout=2):
            pass
    except OSError:
        return None
    opts = Options()
    opts.debugger_address = "127.0.0.1:{}".format(DEBUG_PORT)
    try:
        driver = webdriver.Chrome(service=Service(str(ensure_chromedriver())),
                                  options=opts)
        driver.current_url          # a dead window only shows up on first use
        return driver
    except WebDriverException:
        return None


def shot(driver, tag):
    OUT.mkdir(parents=True, exist_ok=True)
    path = OUT / "_{}.png".format(tag)
    try:
        # ClickPost's show/hide toggle can leave the password rendered as plain text,
        # and a failure screenshot then stores it on disk. Mask the rendering only -
        # the value is untouched, so the form still submits what was typed.
        driver.execute_script(
            "var p=document.querySelector('#input-password');"
            "if(p){p.style.webkitTextSecurity='disc';p.style.textSecurity='disc';}")
    except WebDriverException:
        pass
    try:
        driver.save_screenshot(str(path))
        print("  screenshot: {}".format(path.name))
    except WebDriverException:
        pass


def visible_inputs(driver):
    out = []
    for e in driver.find_elements(By.TAG_NAME, "input"):
        try:
            if e.is_displayed():
                out.append(e)
        except WebDriverException:
            pass
    return out


def button_by_text(driver, *words):
    """The login page carries many type=submit buttons (carousel arrows, links), so a
    button must be matched on its label, never on its type alone."""
    for e in driver.find_elements(By.TAG_NAME, "button"):
        try:
            if not e.is_displayed():
                continue
            label = (e.text or "").strip().lower()
            if not label:
                continue
            for w in words:
                if w == label or w in label.split():
                    return e
        except WebDriverException:
            pass
    return None


def type_into(driver, el, text, label):
    """Replace a field's contents and verify what actually landed there.

    element.clear() is not enough on these React-controlled inputs: the component
    re-applies its own state (and Chrome autofill re-populates), so the typed value
    ends up appended to the old one. Select-all + type replaces the selection through
    real key events, which React does observe. The value is then read back, because a
    silently doubled field surfaces as "Incorrect Username" much later and looks like
    a credentials problem rather than a typing one."""
    for attempt in (1, 2):
        el.click()
        el.send_keys(Keys.CONTROL, "a")
        el.send_keys(Keys.DELETE)
        el.send_keys(text)
        got = el.get_attribute("value") or ""
        if got == text:
            return
        print("    {} did not take (attempt {}) - clearing via JS".format(label, attempt))
        driver.execute_script(
            "arguments[0].value='';"
            "arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", el)
    got = el.get_attribute("value") or ""
    raise RuntimeError("{} field holds {} chars, expected {}".format(
        label, len(got), len(text)))


def wait_enabled(button, timeout=10):
    """The Login button starts styled-disabled and enables once the SPA sees input."""
    end = time.time() + timeout
    while time.time() < end:
        cls = button.get_attribute("class") or ""
        if button.is_enabled() and "cursor-not-allowed" not in cls \
                and "bg-disabled" not in cls:
            return True
        time.sleep(0.3)
    return False


def find_otp_fields(driver):
    """The OTP step may be one field or a row of single-digit boxes - handle both."""
    fields = []
    for e in visible_inputs(driver):
        eid = e.get_attribute("id") or ""
        if eid in ("input-username", "input-password"):
            continue
        if (e.get_attribute("type") or "").lower() == "password":
            continue
        fields.append(e)
    return fields


def describe(driver, tag):
    print("  [{}] url={}".format(tag, driver.current_url))
    for e in visible_inputs(driver):
        print("    input:", {k: e.get_attribute(k) for k in
                             ("type", "name", "id", "placeholder", "maxlength")})
    labels = []
    for e in driver.find_elements(By.TAG_NAME, "button"):
        try:
            if e.is_displayed() and (e.text or "").strip():
                labels.append(e.text.strip().replace("\n", " ")[:24])
        except WebDriverException:
            pass
    print("    buttons:", labels)


def enter_otp(fields, code):
    if len(fields) == 1:
        fields[0].clear()
        fields[0].send_keys(code)
        return
    if len(fields) >= len(code):
        for field, ch in zip(fields, code):
            field.send_keys(ch)          # one character per box
        return
    raise RuntimeError("{} OTP boxes for a {}-character code".format(
        len(fields), len(code)))


def clickpost_login(driver):
    """Run the login flow on `driver`. Returns 0 on success, 1 on failure."""
    require_env(ENV, CLICKPOST_ENV + MAIL_ENV,
                "ClickPost credentials and the OTP mailbox both live in .env.")
    print("ClickPost : {}".format(ENV["CLICKPOST_URL"]))
    print("Username  : {}".format(ENV["CLICKPOST_EMAIL"]))
    print("OTP inbox : {}".format(ENV["OTP_EMAIL"]))

    wait = WebDriverWait(driver, 40)
    try:
        driver.get(ENV["CLICKPOST_URL"])

        # A trusted profile lands straight on the dashboard - no OTP to burn.
        time.sleep(4)
        if "/login" not in driver.current_url:
            print("\nAlready signed in (profile reused) -> {}".format(driver.current_url))
            return 0

        print("\n[1/5] filling credentials")
        user = wait.until(EC.presence_of_element_located(USERNAME_SEL))
        pwd = driver.find_element(*PASSWORD_SEL)
        type_into(driver, user, ENV["CLICKPOST_EMAIL"], "username")
        type_into(driver, pwd, ENV["CLICKPOST_PASSWORD"], "password")

        login_btn = button_by_text(driver, "login", "sign in")
        if login_btn is None:
            describe(driver, "no-login-button")
            raise RuntimeError("Could not find the Login button")
        if not wait_enabled(login_btn):
            print("  warning: Login button still looks disabled - clicking anyway")

        # Baseline BEFORE submitting: anything arriving after this is our OTP.
        print("\n[2/5] snapshotting mailbox")
        baseline = mail_snapshot()

        print("\n[3/5] submitting login")
        login_btn.click()
        time.sleep(6)

        if "/login" not in driver.current_url:
            print("\nLogged in without an OTP step -> {}".format(driver.current_url))
            return 0

        fields = find_otp_fields(driver)
        if not fields:
            shot(driver, "clickpost_no_otp_field")
            raise RuntimeError("Credentials submitted but no OTP field appeared - "
                               "check the screenshot for an error message")
        print("  OTP step detected ({} input field(s))".format(len(fields)))

        print("\n[4/5] waiting for the OTP mail")
        code = wait_for_otp(baseline, timeout=OTP_TIMEOUT)
        print("  OTP received: {}".format(code))

        fields = find_otp_fields(driver) or fields      # re-find, the DOM may redraw
        enter_otp(fields, code)
        time.sleep(1)

        submit = button_by_text(driver, "verify", "submit", "continue", "login", "proceed")
        if submit is not None:
            submit.click()
        else:
            print("  no verify button - assuming the form auto-submits")

        print("\n[5/5] confirming the dashboard loaded")
        try:
            wait.until(lambda d: "/login" not in d.current_url)
        except TimeoutException:
            describe(driver, "otp-not-accepted")
            shot(driver, "clickpost_otp_failed")
            raise RuntimeError("OTP submitted but the page never left /login - "
                               "the code may have been wrong or expired")

        time.sleep(3)
        print("\nLogged in -> {}".format(driver.current_url))
        return 0

    except Exception as e:
        print("\nFAILED: {}: {}".format(e.__class__.__name__, e))
        shot(driver, "clickpost_login_error")
        return 1


def clickpost_session():
    """Return a logged-in driver: reuse the open browser if possible, else log in."""
    driver = attach()
    if driver is not None:
        try:
            # Hand straight to clickpost_login(): it navigates to the login URL and
            # treats a redirect AWAY from it as "already signed in", so it costs one
            # page load when the session is live and logs in properly when it is not.
            #
            # Reading driver.current_url here instead would trust whatever the tab
            # happens to be displaying, which may be a page rendered before the
            # session expired. That produced a run that believed it was signed in,
            # then failed further along with "Could not find 'Bulk Search'" - the
            # session had lapsed and ClickPost had bounced it to /login.
            if clickpost_login(driver) == 0:
                return driver
        except WebDriverException:
            # The window was closed under us; fall through and start a fresh one.
            print("Attached window is gone - starting a new browser")
        driver = None

    driver = start_browser(keep_open=True)
    if clickpost_login(driver) != 0:
        raise RuntimeError("ClickPost login failed")
    return driver


# ============================================================================
# PART 2c - Bulk report
# ============================================================================

ORDERS_URL = "https://dashboard.clickpost.ai/track/order-v2"
REPORTS_URL = "https://dashboard.clickpost.ai/reports?journey=forward"
AWB_LIMIT = 500          # the Bulk Search panel states a 500-AWB cap
DIALOG_SEL = "div.ck-confirmationDialog-wrapper"
# Buttons that continue the task, best first. "Cancel"/"Deselect"/"Discard" are
# deliberately absent - those abandon work rather than proceeding with it.
DIALOG_OK = ["Continue Anyway", "Continue", "Proceed", "Yes", "Confirm", "OK", "Got it"]
PARTIAL_SUFFIXES = (".crdownload", ".tmp", ".partial")


def awbs_from(path):
    from openpyxl import load_workbook
    ws = load_workbook(path)["Logistics Closure Info"]
    header = [c.value for c in ws[1]]
    col = header.index("AWB Number")
    seen, awbs = set(), []
    for row in ws.iter_rows(min_row=2):
        v = row[col].value
        if v is None:
            continue
        v = str(v).strip()
        if v and v not in seen:
            seen.add(v)
            awbs.append(v)
    return awbs


def find_by_text(driver, text, tags=("button", "div", "span", "a", "li", "p")):
    for tag in tags:
        xp = "//{}[normalize-space(text())='{}']".format(tag, text)
        for e in driver.find_elements(By.XPATH, xp):
            try:
                if e.is_displayed():
                    return e
            except WebDriverException:
                pass
    return None


def click_text(driver, text, what=None, timeout=25,
               tags=("button", "div", "span", "a", "li", "p")):
    """Click the first visible element whose exact text matches, waiting for it."""
    what = what or text
    end = time.time() + timeout
    while time.time() < end:
        e = find_by_text(driver, text, tags)
        if e is not None:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", e)
            time.sleep(0.4)
            try:
                e.click()
            except WebDriverException:
                driver.execute_script("arguments[0].click();", e)
            print("  clicked {}".format(what))
            return True
        time.sleep(1)
    raise RuntimeError("Could not find '{}' to click".format(what))


def handle_dialog(driver, timeout=6):
    """Clear any ClickPost confirmation dialog covering the form.

    Re-entering the report builder while a previous unsaved report exists pops a
    dialog which intercepts every click - Selenium then reports the textarea as "not
    clickable" and the real cause stays hidden. The dialog's text and buttons are
    printed before anything is clicked, so an unexpected prompt shows up in the log
    rather than being silently confirmed."""
    end = time.time() + timeout
    while time.time() < end:
        wrappers = [w for w in driver.find_elements(By.CSS_SELECTOR, DIALOG_SEL)
                    if w.is_displayed()]
        if not wrappers:
            return False
        dialog = wrappers[0]
        text = " ".join((dialog.text or "").split())
        labels = []
        for b in dialog.find_elements(By.TAG_NAME, "button"):
            try:
                if b.is_displayed() and (b.text or "").strip():
                    labels.append(b.text.strip())
            except WebDriverException:
                pass
        print("  dialog: {!r} buttons={}".format(text[:110], labels))

        for want in DIALOG_OK:
            for b in dialog.find_elements(By.TAG_NAME, "button"):
                try:
                    if b.is_displayed() and (b.text or "").strip().lower() == want.lower():
                        try:
                            b.click()
                        except WebDriverException:
                            driver.execute_script("arguments[0].click();", b)
                        print("  dialog: clicked '{}'".format(want))
                        time.sleep(2)
                        return True
                except WebDriverException:
                    pass
        raise RuntimeError("Unrecognised dialog blocking the page: {!r} buttons={}"
                           .format(text[:150], labels))
    return False


def particulars_box(driver, timeout=25):
    end = time.time() + timeout
    while time.time() < end:
        for e in driver.find_elements(By.TAG_NAME, "textarea"):
            try:
                ph = (e.get_attribute("placeholder") or "").lower()
                if e.is_displayed() and "particular" in ph:
                    return e
            except WebDriverException:
                pass
        time.sleep(1)
    raise RuntimeError("Particulars box never appeared")


def fill_particulars(box, awbs):
    """Type the AWBs one per line, then read the value back.

    Read-back matters: this is a React-controlled textarea, and a value that does not
    stick produces an empty report rather than an error."""
    text = "\n".join(awbs)
    box.click()
    box.send_keys(Keys.CONTROL, "a")
    box.send_keys(Keys.DELETE)
    box.send_keys(text)
    time.sleep(1)
    got = (box.get_attribute("value") or "").strip()
    lines = [x for x in got.splitlines() if x.strip()]
    if len(lines) != len(awbs):
        raise RuntimeError("Particulars holds {} line(s), expected {}".format(
            len(lines), len(awbs)))
    print("  particulars: {} AWB(s) entered".format(len(lines)))


def field_checkboxes(driver):
    """Every visible tickable control on the Fields tab."""
    out = []
    for e in driver.find_elements(By.CSS_SELECTOR, "input[type=checkbox]"):
        try:
            if e.is_displayed():
                out.append(("input", e))
        except WebDriverException:
            pass
    # The UI also uses Material icon spans in place of real inputs in some places.
    for e in driver.find_elements(By.XPATH, "//span[text()='check_box_outline_blank']"):
        try:
            if e.is_displayed():
                out.append(("icon", e))
        except WebDriverException:
            pass
    return out


def select_all_fields(driver):
    """Tick every column. Prefers a real select-all control, else ticks each box."""
    for label in ("Select All", "Select all", "All Fields", "All"):
        e = find_by_text(driver, label)
        if e is not None:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", e)
            time.sleep(0.3)
            try:
                e.click()
            except WebDriverException:
                driver.execute_script("arguments[0].click();", e)
            print("  used '{}' control".format(label))
            time.sleep(2)
            return "select-all:" + label

    boxes = field_checkboxes(driver)
    print("  no select-all control - ticking {} box(es) individually".format(len(boxes)))
    ticked = 0
    for kind, e in boxes:
        try:
            if kind == "input" and e.is_selected():
                continue
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", e)
            try:
                e.click()
            except WebDriverException:
                driver.execute_script("arguments[0].click();", e)
            ticked += 1
            time.sleep(0.15)
        except WebDriverException:
            pass
    print("  ticked {} field(s)".format(ticked))
    return "individual:{}".format(ticked)


def run_bulk_report(driver, tickets_path, dry=False):
    """Orders -> Bulk Search -> Open Reports -> AWBs -> all Fields -> Generate."""
    awbs = awbs_from(tickets_path)
    if not awbs:
        print("{} has no AWB numbers - nothing to search".format(tickets_path.name))
        return False
    if len(awbs) > AWB_LIMIT:
        raise SystemExit("{} AWBs exceeds the {} cap".format(len(awbs), AWB_LIMIT))
    print("Source : {}".format(tickets_path.name))
    print("AWBs   : {} -> {}".format(len(awbs), ", ".join(awbs)))

    print("\n[1/6] opening Orders")
    driver.get(ORDERS_URL)
    time.sleep(6)

    print("\n[2/6] Bulk Search")
    click_text(driver, "Bulk Search", tags=("div", "span", "button", "a"))
    time.sleep(3)

    print("\n[3/6] Open Reports")
    click_text(driver, "Open Reports", tags=("span", "div", "button", "a"))
    time.sleep(6)
    # Wait on the textarea, not on the "Particulars" caption: that caption is a
    # <label> whose text node is split around the red asterisk, so matching it
    # reports "screen did not open" while the screen is plainly open.
    try:
        box = particulars_box(driver, timeout=30)
    except RuntimeError:
        shot(driver, "bulk_no_configure")
        raise RuntimeError("Configure Report screen did not open")

    print("\n[4/6] entering AWBs")
    handle_dialog(driver)
    awb_radio = find_by_text(driver, "AWB", tags=("p", "span", "div", "label"))
    if awb_radio is not None:
        try:
            awb_radio.click()          # AWB is the default; this is a safety net
        except WebDriverException:
            pass
    fill_particulars(box, awbs)

    print("\n[5/6] Fields - selecting all columns")
    click_text(driver, "Fields", tags=("div", "span", "button", "li"))
    time.sleep(4)
    how = select_all_fields(driver)
    time.sleep(2)
    # Ticking every column selects PII fields, and ClickPost then blocks the screen
    # with "Sensitive Fields Selected" (Drop Phone No. is encrypted, so generation
    # runs slower). Continue Anyway keeps all columns, which is what is wanted;
    # Deselect would silently drop them.
    handle_dialog(driver)
    selected = ""
    for e in driver.find_elements(By.XPATH, "//*[contains(text(),' selected')]"):
        try:
            if e.is_displayed():
                selected = (e.text or "").strip()
                break
        except WebDriverException:
            pass
    print("  fields tab reports: {}".format(selected or "(count not shown)"))

    print("\n[6/6] Generate Report")
    if dry:
        print("  --dry: stopping before the click")
        return False
    click_text(driver, "Generate Report", tags=("div", "span", "button"))
    time.sleep(6)
    handle_dialog(driver)              # it can also fire on submit
    time.sleep(4)
    print("\nGenerate Report clicked ({}).".format(how))
    return True


# ============================================================================
# PART 2d - Download the finished report
# ============================================================================


def read_report_rows(driver):
    """Report rows as dicts, newest first (the table is already newest-first)."""
    rows = []
    for tr in driver.find_elements(By.TAG_NAME, "tr"):
        try:
            cells = [td.text.strip() for td in tr.find_elements(By.TAG_NAME, "td")]
        except WebDriverException:
            continue
        if len(cells) >= 6 and cells[0].isdigit():
            rows.append({"id": cells[0], "name": cells[1], "data": cells[2],
                         "status": cells[4], "requested": cells[5], "el": tr})
    return rows


def run_download(driver, wanted_id=None, timeout=180):
    """Download a finished report into output/. Returns the saved path."""
    OUT.mkdir(parents=True, exist_ok=True)
    before = {p.name for p in OUT.iterdir()}

    # Set over CDP, not through Chrome options: this driver may be attached to a
    # browser it did not start, and options only apply at launch.
    driver.execute_cdp_cmd("Page.setDownloadBehavior",
                           {"behavior": "allow", "downloadPath": str(OUT)})
    driver.get(REPORTS_URL)
    time.sleep(8)

    rows = read_report_rows(driver)
    if not rows:
        raise RuntimeError("No report rows found on {}".format(driver.current_url))

    print("Reports listed:")
    for r in rows[:5]:
        print("  {}  {:<10} {}".format(r["id"], r["status"], r["requested"]))

    row = next((r for r in rows if r["id"] == wanted_id), None) if wanted_id else rows[0]
    if row is None:
        raise RuntimeError("Report {} is not in the list".format(wanted_id))
    if row["status"].lower() != "success":
        raise RuntimeError("Report {} is '{}', not Success - nothing to download yet"
                           .format(row["id"], row["status"]))
    print("\nDownloading report {} ({}, requested {})".format(
        row["id"], row["data"], row["requested"]))

    icon = None
    for e in row["el"].find_elements(By.XPATH,
                                     ".//*[normalize-space(text())='file_download']"):
        if e.is_displayed():
            icon = e
            break
    if icon is None:
        raise RuntimeError("No download control in row {}".format(row["id"]))
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", icon)
    time.sleep(0.5)
    try:
        icon.click()
    except WebDriverException:
        driver.execute_script("arguments[0].click();", icon)

    # Wait for a genuinely new, fully-written file.
    deadline = time.time() + timeout
    while time.time() < deadline:
        fresh = [p for p in OUT.iterdir()
                 if p.name not in before and not p.name.endswith(PARTIAL_SUFFIXES)
                 and p.suffix.lower() in (".csv", ".xlsx")]
        if fresh:
            got = max(fresh, key=lambda p: p.stat().st_mtime)
            size = got.stat().st_size
            time.sleep(2)
            if got.stat().st_size == size and size > 0:
                print("\nSaved: {}".format(got))
                print("Size : {:,} bytes".format(size))
                return got
        time.sleep(2)

    raise RuntimeError("No file appeared in {} within {}s".format(OUT, timeout))


# ============================================================================
# PART 3 - Merge and Mapping
# ============================================================================

SRC_SHEET = "Logistics Closure Info"
STATUS_COL = "Latest Status"
AWB_TICKETS = "AWB Number"
AWB_CLICKPOST = "AWB"
# Delivered is a finished shipment; anything RTO-prefixed (RTO, RTOInTransit,
# RTODelivered, ...) is already on its way back. Neither needs chasing.
EXCLUDED_EXACT = {"delivered"}
EXCLUDED_PREFIXES = ("rto",)

# (source column, output column) for Mapping.xlsx
MAPPING_COLUMNS = [
    ("Ticket Number", "Ticket Number"),
    ("AWB Number", "AWB Number"),
    ("Vinculum Shipment EDD", "Vinc Shipment EDD"),
    (None, "Delay Days"),                       # computed
    ("Logistics Classification", "Concern Type"),
    ("Courier Partner", "Courier Partner"),
    ("States", "State"),
]
MAPPING_WIDTHS = [15, 18, 20, 12, 24, 18, 18]
# Derived, not hard-coded: the sort and the date format both key off the EDD column,
# and inserting a column ahead of it silently shifted those when they were literals.
EDD_OUT_IDX = [out for _, out in MAPPING_COLUMNS].index("Vinc Shipment EDD")
DATE_FORMATS = ["%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%b-%Y", "%b %d, %Y"]


def norm_status(v):
    return "".join(str(v or "").split()).lower()


def is_excluded_status(status):
    n = norm_status(status)
    return n in EXCLUDED_EXACT or n.startswith(EXCLUDED_PREFIXES)


def parse_edd(v):
    """Return a date, or None when the cell is empty or unparseable."""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    s = re.split(r"[ T]", s)[0]          # drop any time part
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def load_clickpost_statuses(path):
    """AWB -> Latest Status. The first row for an AWB wins: the report is ordered
    newest-first, so that row is the current one."""
    with path.open(encoding="utf-8-sig", newline="") as fh:
        reader = csv.DictReader(fh)
        if STATUS_COL not in (reader.fieldnames or []):
            raise RuntimeError("{} has no '{}' column".format(path.name, STATUS_COL))
        out = {}
        for row in reader:
            awb = norm_awb(row.get(AWB_CLICKPOST))
            if awb and awb not in out:
                out[awb] = (row.get(STATUS_COL) or "").strip()
    return out


def run_merge(tickets_path, cp_path):
    """Add Latest Status to the ticket export, dropping Delivered/RTO/not-found.

    The source workbook is never modified - the merged sheet is written alongside it
    as *_with_status.xlsx, so a bad merge cannot cost a day of ticket data."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    print("Tickets   : {}".format(tickets_path.name))
    print("ClickPost : {}".format(cp_path.name))
    statuses = load_clickpost_statuses(cp_path)
    print("ClickPost AWBs: {}".format(len(statuses)))

    ws_in = load_workbook(tickets_path)[SRC_SHEET]
    header = [c.value for c in ws_in[1]]
    if AWB_TICKETS not in header:
        raise RuntimeError("{} has no '{}' column".format(
            tickets_path.name, AWB_TICKETS))
    awb_idx = header.index(AWB_TICKETS)

    kept, dropped = [], []
    for row in ws_in.iter_rows(min_row=2):
        values = [c.value for c in row]
        if all(v is None for v in values):
            continue
        awb = norm_awb(values[awb_idx])
        status = statuses.get(awb)
        if status is None:
            dropped.append((awb, "", "not in ClickPost"))
        elif is_excluded_status(status):
            dropped.append((awb, status, "excluded status"))
        else:
            kept.append(values + [status])

    print("\nper-ticket outcome:")
    for awb, status, why in dropped:
        print("  DROP {:<14} {:<16} {}".format(awb, status or "-", why))
    for values in kept:
        print("  KEEP {:<14} {}".format(norm_awb(values[awb_idx]), values[-1]))

    if not kept:
        print("\nNothing survived the filter - no file written.")
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = SRC_SHEET
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E79")
    ws.append(header + [STATUS_COL])
    for c in ws[1]:
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for values in kept:
        ws.append(values)

    widths = [15, 20, 46, 14, 22, 30, 26, 18, 18, 22, 12, 16, 24]
    for i, w in enumerate(widths[:len(header)] + [18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(
        get_column_letter(len(header) + 1), ws.max_row)

    info = wb.create_sheet("Merge Info")
    info.append(["Generated (IST)", datetime.now(IST).strftime("%Y-%m-%d %H:%M:%S")])
    info.append(["Ticket source", tickets_path.name])
    info.append(["ClickPost source", cp_path.name])
    info.append(["Tickets in", len(kept) + len(dropped)])
    info.append(["Kept", len(kept)])
    info.append(["Dropped - excluded status",
                 sum(1 for d in dropped if d[2] == "excluded status")])
    info.append(["Dropped - not in ClickPost",
                 sum(1 for d in dropped if d[2] == "not in ClickPost")])
    info.append([])
    info.append(["Dropped AWB", "Status", "Reason"])
    for awb, status, why in dropped:
        info.append([awb, status, why])
    info.column_dimensions["A"].width = 26
    info.column_dimensions["B"].width = 20
    info.column_dimensions["C"].width = 22
    for c in info["A"]:
        c.font = Font(bold=True)

    out_path = tickets_path.with_name(tickets_path.stem + "_with_status.xlsx")
    wb.save(out_path)
    print("\nSaved: {}".format(out_path))
    print("Rows : {} kept, {} dropped".format(len(kept), len(dropped)))
    return out_path


def run_mapping(src, asof=None):
    """Build Mapping.xlsx - the trimmed follow-up sheet.

    Delay Days = today (IST) - Vinc Shipment EDD, in whole days. Two cases the source
    data actually produces, both handled deliberately rather than silently:
      * EDD blank  -> Delay Days blank. Writing 0 would read as "due today, on time",
                      a different claim from "we were never given a date".
      * EDD future -> a negative number, i.e. days still to run. Not clamped to 0, so
                      a not-yet-due shipment stays distinguishable from one due today."""
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    today = parse_edd(asof) if asof else datetime.now(IST).date()
    if today is None:
        raise SystemExit("Could not read --asof {!r}".format(asof))
    print("Source : {}".format(src.name))
    print("As of  : {} (IST)".format(today))

    ws_in = load_workbook(src)[SRC_SHEET]
    header = [c.value for c in ws_in[1]]
    missing = [s for s, _ in MAPPING_COLUMNS if s and s not in header]
    if missing:
        raise RuntimeError("{} is missing column(s): {}".format(
            src.name, ", ".join(missing)))
    idx = {s: header.index(s) for s, _ in MAPPING_COLUMNS if s}

    rows, no_edd = [], []
    for row in ws_in.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        edd = parse_edd(row[idx["Vinculum Shipment EDD"]])
        awb = row[idx["AWB Number"]]
        if edd is None:
            no_edd.append(awb)
        rows.append([
            row[idx["Ticket Number"]],
            awb,
            edd if edd else "",
            (today - edd).days if edd else "",
            row[idx["Logistics Classification"]],
            row[idx["Courier Partner"]],
            row[idx["States"]],
        ])

    # Most delayed first - oldest EDD at the top, exactly as the reference report is
    # read. Rows with no EDD have no delay to rank on, so they sit at the bottom
    # rather than being treated as either very old or brand new.
    rows.sort(key=lambda r: (r[EDD_OUT_IDX] == "",
                             r[EDD_OUT_IDX] if r[EDD_OUT_IDX] != "" else date.min))

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapping"
    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="1F4E79")
    ws.append([out for _, out in MAPPING_COLUMNS])
    for c in ws[1]:
        c.font, c.fill = head_font, head_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    edd_letter = get_column_letter(EDD_OUT_IDX + 1)
    for cell in ws[edd_letter][1:]:
        cell.number_format = "DD-MMM-YYYY"     # 22-Jul-2026, matching the mail body
    for i, w in enumerate(MAPPING_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:{}{}".format(
        get_column_letter(len(MAPPING_COLUMNS)), ws.max_row)

    path = OUT / "Mapping.xlsx"
    wb.save(path)

    # Print whatever MAPPING_COLUMNS says, so adding a column cannot leave the
    # console showing a different set of fields from the file.
    out_names = [out for _, out in MAPPING_COLUMNS]
    shown = [[str(c) if c != "" else "" for c in r] for r in rows]
    widths = [max([len(out_names[i])] + [len(r[i]) for r in shown]) + 2
              for i in range(len(out_names))]
    print()
    print("".join(h.ljust(widths[i]) for i, h in enumerate(out_names)).rstrip())
    for r in shown:
        print("".join(c.ljust(widths[i]) for i, c in enumerate(r)).rstrip())
    if no_edd:
        print("\nNote: no EDD on {} - Delay Days left blank".format(
            ", ".join(map(str, no_edd))))
    print("\nSaved: {}  ({} row(s))".format(path, len(rows)))
    return path, len(rows)


# ============================================================================
# PART 4 - Mail the Mapping file
# ============================================================================

SEND_ENV = ["MAIL_TO", "SMTP_HOST", "SMTP_PORT", "OTP_EMAIL", "OTP_EMAIL_PASSWORD"]

# Header and footer are fixed wording. Only the addressee changes, via
# MAIL_GREETING_NAME in .env, so renaming the recipient needs no code edit.
MAIL_INTRO = (
    "Please look into the below aged shipments on priority and arrange delivery "
    "at the earliest.")
# (plain text, html) pairs - the html half bolds the same phrases as the reference
# screenshot: the concern types, and the deadline in the footer.
MAIL_INTRO_2 = (
    "Several shipments are already significantly beyond the EDD, with "
    "Out TAT Intransit/Fake NDR concerns.",
    "Several shipments are already significantly beyond the EDD, with "
    "<b>Out TAT Intransit/Fake NDR</b> concerns.")
MAIL_FOOTER = (
    "Please treat this as urgent and ensure resolution within 24 working hours.",
    "Please treat this as urgent and ensure resolution within "
    "<b>24 working hours</b>.")
MAIL_DATE_FMT = "%d-%b-%Y"          # 22-Jul-2026, as in the reference screenshot
MAIL_SUBJECT = "Customer Escalation - Pending Cases"
# Subject date is dd-mm-yy ("19-08-26"), deliberately NOT the dd-MMM-yyyy used in the
# table. The fixed prefix plus this date is the human-readable handle on a thread, and
# a secondary way to find a reply when the In-Reply-To header is missing or rewritten.
SUBJECT_DATE_FMT = "%d-%m-%y"


def read_mapping(path):
    """Read Mapping.xlsx back for the mail body.

    The table is rendered from the SAVED FILE, not from in-memory values, so the
    mail cannot drift from the attachment: what is quoted is by construction what
    was written. Dates are rendered yyyy-mm-dd to match the sheet's number format,
    and blanks stay blank - an empty Delay Days means no EDD was set, which is not
    the same claim as zero."""
    from openpyxl import load_workbook
    ws = load_workbook(path)["Mapping"]
    header = [str(c.value) for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None or v == "" for v in r):
            continue
        out = []
        for v in r:
            if v is None or v == "":
                out.append("")
            elif isinstance(v, datetime):
                out.append(v.strftime(MAIL_DATE_FMT))
            elif isinstance(v, date):
                out.append(v.strftime(MAIL_DATE_FMT))
            else:
                out.append(str(v))
        rows.append(out)
    return header, rows


def _esc(s):
    return (str(s).replace("&", "&amp;").replace("<", "&lt;")
            .replace(">", "&gt;").replace('"', "&quot;"))


def table_html(header, rows):
    """Bordered, centered, navy-header table matching the reference screenshot."""
    th = "".join(
        '<th style="border:1px solid #1F3864;background:#2F5496;color:#ffffff;'
        'padding:6px 12px;text-align:center;font-weight:bold;">{}</th>'.format(_esc(h))
        for h in header)
    trs = []
    for r in rows:
        tds = "".join(
            '<td style="border:1px solid #1F3864;padding:5px 12px;'
            'text-align:center;">{}</td>'.format(_esc(c)) for c in r)
        trs.append("<tr>{}</tr>".format(tds))
    return ('<table cellspacing="0" cellpadding="0" '
            'style="border-collapse:collapse;background:#ffffff;'
            'font-family:Segoe UI,Arial,sans-serif;font-size:13px;color:#000000;">'
            "<thead><tr>{}</tr></thead><tbody>{}</tbody></table>").format(
                th, "".join(trs))


def table_text(header, rows):
    widths = [len(h) for h in header]
    for r in rows:
        for i, c in enumerate(r):
            widths[i] = max(widths[i], len(c))
    line = "  ".join(h.ljust(widths[i]) for i, h in enumerate(header)).rstrip()
    out = [line, "-" * len(line)]
    for r in rows:
        out.append("  ".join(c.ljust(widths[i]) for i, c in enumerate(r)).rstrip())
    return "\n".join(out)


def send_mapping_mail(path, rows):
    """Mail Mapping.xlsx. Returns True only if SMTP accepted the message.

    This is the commit point for the whole pipeline: while it returns False the
    watermark stays put and the same tickets are retried on the next run."""
    import smtplib
    from email.message import EmailMessage

    require_env(ENV, SEND_ENV,
                "MAIL_TO is the report recipient; SMTP_* and OTP_EMAIL* are the "
                "Zoho mailbox used to send.")
    sender = (ENV.get("MAIL_FROM") or "").strip() or ENV["OTP_EMAIL"]
    to = [a.strip() for a in ENV["MAIL_TO"].split(",") if a.strip()]
    cc = [a.strip() for a in (ENV.get("MAIL_CC") or "").split(",") if a.strip()]
    today = datetime.now(IST).strftime(MAIL_DATE_FMT)

    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = ", ".join(to)
    if cc:
        msg["Cc"] = ", ".join(cc)
    msg["Subject"] = "{} | {}".format(
        MAIL_SUBJECT, datetime.now(IST).strftime(SUBJECT_DATE_FMT))
    # Generated here rather than left to the SMTP server: this id is how replies are
    # later matched back to their thread, so we have to know it.
    msg["Message-ID"] = make_msgid(domain=sender.split("@")[-1])

    greeting = "Hi {},".format((ENV.get("MAIL_GREETING_NAME") or "Rahul").strip())
    header, table_rows = read_mapping(path)

    msg.set_content("{}\n\n{}\n\n{}\n\n{}\n\n{}\n".format(
        greeting, MAIL_INTRO, MAIL_INTRO_2[0],
        table_text(header, table_rows), MAIL_FOOTER[0]))
    msg.add_alternative(
        '<div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;'
        'color:#000000;">'
        "<p>{}</p><p>{}</p><p>{}</p>{}<p>{}</p></div>".format(
            _esc(greeting), _esc(MAIL_INTRO), MAIL_INTRO_2[1],
            table_html(header, table_rows), MAIL_FOOTER[1]),
        subtype="html")
    msg.add_attachment(
        path.read_bytes(),
        maintype="application",
        subtype="vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=path.name)

    print("From    : {}".format(sender))
    print("To      : {}{}".format(", ".join(to),
                                  "  cc " + ", ".join(cc) if cc else ""))
    print("Subject : {}".format(msg["Subject"]))
    try:
        with smtplib.SMTP(ENV["SMTP_HOST"], int(ENV["SMTP_PORT"]), timeout=60) as s:
            s.starttls()
            s.login(ENV["OTP_EMAIL"], ENV["OTP_EMAIL_PASSWORD"])
            refused = s.send_message(msg, from_addr=sender, to_addrs=to + cc)
        if refused:
            # A partial refusal still means somebody did not get the report.
            print("FAILED: server refused {}".format(refused))
            return None
        print("Mail sent.")
        return msg["Message-ID"]          # the thread handle for reply tracking
    except Exception as e:
        print("FAILED to send mail: {}: {}".format(e.__class__.__name__, e))
        return None


# ============================================================================
# PART 5 - Replies: watch our own threads, comment statuses onto tickets
# ============================================================================
#
# Only threads this pipeline created are ever touched. Each sent report is recorded
# with the Message-ID we generated for it, and a reply is matched by that id coming
# back in In-Reply-To / References. The dated subject is a fallback for clients that
# rewrite headers. Nothing else in the mailbox is read.
#
# The reply is the same table sent back with a Status column appended. A reply also
# QUOTES the original mail, so the HTML holds two tables listing the same AWBs - the
# quoted one has no Status column. Only a table whose header carries Status is read;
# without that rule every ticket gets an empty status commented onto it.

FOLLOWUP_HOUR = 15            # 15:00 IST the day after, if still no reply
FOLLOWUP_TEXT = "Hi, please revert as soon as possible."
COMMENT_PATH = "tickets/{}/comments"

REPLY_FOLLOWUP_HOURS = 15     # chase again this long after a reply lands
FOLLOWUP_MAX_ROUNDS = 10      # ~6 days of chasing, then leave the thread alone

# "Undelivered" and "Not delivered" both CONTAIN "delivered". A plain substring
# test would read either as final and abandon a live shipment forever, so the
# negations are matched first and win. The gap allows "not yet delivered" and
# "could not be delivered" without swallowing a whole sentence.
_UNDELIVERED = re.compile(r"\bun-?delivered\b")
_NEGATED = re.compile(r"\b(not|never|no|cannot|couldn'?t|could not|unable|fail\w*)"
                      r"[\w\s,'-]{0,24}\bdelivered\b")


def is_final_status(text):
    """True when this remark means the AWB needs no further chasing.

    Final is "delivered" anywhere in the remark, minus negations, plus anything
    RTO-prefixed - which matches the ClickPost filter's reasoning that a shipment
    already heading back needs no chasing either."""
    t = " ".join(str(text or "").split()).lower()
    if not t:
        return False
    if _UNDELIVERED.search(t):
        return False
    if _NEGATED.search(t):
        return False
    if "delivered" in t:
        return True
    return t.replace(" ", "").startswith("rto")


def is_delivered_status(text):
    """True when the remark says the shipment was actually delivered.

    Persistence only - this decides the Delivered Date and nothing else. It is
    deliberately NARROWER than is_final_status, which answers "stop chasing" and
    is therefore true of plain RTO as well: a parcel heading back needs no chase,
    but nothing has been delivered to anyone yet.

    "RTO Delivered" DOES count - the shipment was delivered, back to origin - and
    falls out of the same substring test without a special case. The negations
    are is_final_status's own, so "Undelivered" and "could not be delivered" can
    never stamp a Delivered Date."""
    t = " ".join(str(text or "").split()).lower()
    if not t or _UNDELIVERED.search(t) or _NEGATED.search(t):
        return False
    return "delivered" in t


def thread_ticket_map(mapping_path):
    """(header, {ticketNumber: {"ticketId", "awb", "row"}}) for the mail just sent.

    Keyed on the ticket number because that is the unique identifier - the same AWB
    can carry more than one ticket, so an AWB-keyed map would silently lose one.

    Recorded on the thread itself because Mapping.xlsx is deleted once the run
    commits: by the time a reply arrives days later the file is long gone. The
    internal ticket id comes from the AWB registry, which captured it at fetch time."""
    header, rows = read_mapping(mapping_path)
    try:
        awb_i = header.index("AWB Number")
        tkt_i = header.index("Ticket Number")
    except ValueError:
        return {}
    registry = load_registry()
    out = {}
    for r in rows:
        num = str(r[tkt_i] or "").strip()
        awb = norm_awb(r[awb_i])
        if not num:
            continue
        # The registry is only trusted when its entry is for THIS ticket number.
        # It is keyed by AWB, and an AWB can carry several tickets, so taking the id
        # from an AWB hit alone could bind this row to a different ticket entirely.
        entry = registry.get(awb) or {}
        tid = (entry.get("ticketId")
               if str(entry.get("ticketNumber") or "") == num else None)
        if not tid:
            tid = resolve_ticket_id(num)
            if tid and awb and str((registry.get(awb) or {}).get("ticketNumber") or "") == num:
                registry[awb]["ticketId"] = tid
                save_registry(registry)
        if not tid:
            print("  warning: no ticket id for #{} - a reply for it cannot be "
                  "commented".format(num))
        # The whole row is kept, not just the ids: Mapping.xlsx is deleted when
        # the run commits, so this is the only surviving copy of the courier,
        # EDD and concern columns - and the chase re-sends them unchanged.
        out[num] = {"ticketId": tid, "awb": awb, "row": list(r)}
        cells = dict(zip(header, r))
        db.upsert_ticket(num, ticket_id=tid, awb=awb, extra={
            "courier": cells.get("Courier Partner"),
            "logistics_class": cells.get("Concern Type"),
            "states": cells.get("State"),
        })
        db.register_awb(awb, num)
    return header, out


def record_thread(message_id, subject, ticket_map, columns=None):
    # Locked: the mail has already gone out, so losing this record to a
    # concurrent reply scan would mail the courier the same escalation again.
    with threads_lock(label="record_thread"):
        return _record_thread(message_id, subject, ticket_map, columns)


def _record_thread(message_id, subject, ticket_map, columns=None):
    threads = load_threads()
    sent_at = datetime.now(IST)
    # Mirrored to Postgres as the mail goes out, so the database is a live
    # record rather than something a migration has to catch up on later.
    email_id = db.record_email(
        message_id=message_id, subject=subject, sent_at=sent_at,
        ticket_map=ticket_map, columns=columns, kind="escalation",
        from_addr=(ENV.get("MAIL_FROM") or ENV.get("OTP_EMAIL")),
        to_addrs=[a.strip() for a in (ENV.get("MAIL_TO") or "").split(",")
                  if a.strip()],
        cc_addrs=[a.strip() for a in (ENV.get("MAIL_CC") or "").split(",")
                  if a.strip()])
    threads.append({
        "message_id": message_id,
        "subject": subject,
        "db_email_id": email_id,
        "sent_ist": sent_at.isoformat(),
        "tickets": ticket_map,
        # The mapping header, so a chase can rebuild the same table we escalated
        # with rather than a reduced one.
        "columns": list(columns or []),
        "status": "awaiting_reply",
        "followup_sent_ist": None,
        "processed": {},
        "seen_replies": [],
    })
    save_threads(threads)
    print("Thread recorded: {} ticket(s) awaiting reply{}".format(
        len(ticket_map),
        "  (db id {})".format(email_id) if email_id else ""))


def load_threads():
    if THREADS_FILE.exists():
        return json.loads(THREADS_FILE.read_text(encoding="utf-8")).get("threads", [])
    return []


def save_threads(threads):
    """Atomic replace, so a concurrent reader never sees a half-written file.

    write_text truncates first and then writes, so a reader landing in between got
    truncated JSON. os.replace swaps the finished file in as one operation."""
    tmp = THREADS_FILE.with_name(THREADS_FILE.name + ".tmp")
    tmp.write_text(json.dumps({"threads": threads}, indent=2), encoding="utf-8")
    os.replace(tmp, THREADS_FILE)


THREADS_LOCK_FILE = BASE / ".threads.lock"


@contextlib.contextmanager
def threads_lock(timeout=300, poll=0.25, label=""):
    """Exclusive cross-process lock over threads.json. Yields True if it was taken.

    WHY: everything that touches threads.json does load -> work -> save on a WHOLE-FILE
    in-memory copy. A reply scan holds that copy open for as long as an IMAP fetch per
    open thread plus a Desk POST takes - seconds to minutes. If the daily run's
    record_thread() appends inside that window, the scan then writes its stale list
    back over it and the new record is gone. The escalation has ALREADY been mailed by
    then, so it is unrecoverable: the reply is never matched to a ticket, no chase ever
    fires, and reported_tickets() stops suppressing those ticket numbers - so the
    courier is sent the same escalation a second time, which is the exact failure
    4094fd7 was written to prevent.

    Everything that read-modify-writes the file takes this lock: the reply scan, the
    daily run's record_thread(), and the follow-up sweep - plus webhook_server.py,
    which drives the scan from a separate process.

    On timeout it proceeds UNLOCKED rather than aborting. A lost comment is picked up
    by the next scan; a daily pipeline that refuses to record a mail it has already
    sent is worse than the race it is avoiding."""
    THREADS_LOCK_FILE.touch(exist_ok=True)
    f = open(THREADS_LOCK_FILE, "a+b")
    deadline = time.time() + timeout
    locked = False
    try:
        while True:
            try:
                f.seek(0)
                if IS_WINDOWS:
                    import msvcrt
                    msvcrt.locking(f.fileno(), msvcrt.LK_NBLCK, 1)
                else:
                    import fcntl
                    fcntl.flock(f.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                locked = True
                break
            except OSError:
                if time.time() >= deadline:
                    print("  warning: threads.json lock held by another process for "
                          "{}s - proceeding without it{}".format(
                              timeout, " ({})".format(label) if label else ""))
                    break
                time.sleep(poll)
        yield locked
    finally:
        if locked:
            try:
                f.seek(0)
                if IS_WINDOWS:
                    import msvcrt
                    msvcrt.locking(f.fileno(), msvcrt.LK_UNLCK, 1)
                else:
                    import fcntl
                    fcntl.flock(f.fileno(), fcntl.LOCK_UN)
            except OSError:
                pass
        f.close()


def api_post(path, payload, retries=3):
    """POST to Desk with the same token refresh / backoff as api_get."""
    global TOKEN
    body = json.dumps(payload).encode("utf-8")
    for attempt in range(retries):
        headers = {"Authorization": "Zoho-oauthtoken " + TOKEN,
                   "Content-Type": "application/json"}
        org = (ENV.get("ZOHO_ORG_ID") or "").strip()
        if org and org != PLACEHOLDER:
            headers["orgId"] = org
        req = urllib.request.Request(api_base() + path, data=body, headers=headers,
                                     method="POST")
        try:
            with urllib.request.urlopen(req, timeout=45) as r:
                return json.load(r)
        except urllib.error.HTTPError as e:
            if e.code == 401 and attempt < retries - 1:
                TOKEN = get_token()
                continue
            if e.code in (429, 500, 502, 503) and attempt < retries - 1:
                time.sleep(2 ** attempt)
                continue
            raise RuntimeError("HTTP {} from {}: {}".format(
                e.code, path, e.read()[:300].decode("utf-8", "ignore")))
        except (urllib.error.URLError, TimeoutError):
            if attempt < retries - 1:
                time.sleep(2 ** attempt)
                continue
            raise
    raise RuntimeError("unreachable")


def cell_text(html_fragment):
    s = re.sub(r"(?s)<[^>]+>", " ", html_fragment)
    s = (s.replace("&nbsp;", " ").replace("&amp;", "&").replace("&lt;", "<")
          .replace("&gt;", ">").replace("&#39;", "'").replace("\xa0", " "))
    return re.sub(r"\s+", " ", s).strip()


SENT_HEADERS = {re.sub(r"[^a-z]", "", out.lower()) for _, out in MAPPING_COLUMNS}


def _looks_like_status(label):
    """Is this header cell the reply's status column?

    Not an equality test on purpose. A real recipient types this header by hand, and
    the first live reply already came back as "Sttaus". Close spellings and the
    obvious synonyms count."""
    s = re.sub(r"[^a-z]", "", label.lower())
    if not s:
        return False
    if s.startswith(("status", "remark", "update", "comment", "reply")):
        return True
    import difflib
    return difflib.SequenceMatcher(None, s, "status").ratio() >= 0.7


def _status_index(cells):
    """Index of the status column in a header row, or None.

    Falls back to structure when the label is unrecognisable: the reply is our own
    table with one extra column on the end, so a header whose leading cells are all
    ours plus a trailing extra means that extra IS the status, whatever it is called."""
    known = [re.sub(r"[^a-z]", "", c.lower()) in SENT_HEADERS for c in cells]
    for i, c in enumerate(cells):
        # A column we sent is never the status column, however similar the word
        # looks: "State" scores 0.73 against "status" and would otherwise win.
        if not known[i] and _looks_like_status(c):
            return i
    if len(cells) >= 2 and all(known[:-1]) and not known[-1] and cells[-1].strip():
        return len(cells) - 1
    return None


def parse_status_table(html):
    """Rows of {"ticket", "awb", "status"} from the reply's table.

    Both identifiers are returned because they are not equivalent: a ticket number is
    unique, an AWB is not - one shipment can have several tickets raised against it.
    The ticket column is preferred downstream; the AWB is the fallback for replies
    that don't carry it (Bluedart's own format has no ticket column).

    Walks rows in document order. A header row naming AWB or Ticket turns parsing on
    only if it also carries a status column; one without (the quoted original) turns
    it back off - otherwise the quoted copy of our own mail would blank every ticket."""
    rows = []
    active = False
    awb_i = ticket_i = status_i = None
    for tr in re.findall(r"(?is)<tr[^>]*>(.*?)</tr>", html or ""):
        cells = [cell_text(c) for c in
                 re.findall(r"(?is)<t[dh][^>]*>(.*?)</t[dh]>", tr)]
        if not any(cells):
            continue
        low = [c.lower() for c in cells]
        if any("awb" in c for c in low) or any("ticket" in c for c in low):
            si = _status_index(cells)
            if si is not None:
                awb_i = next((i for i, c in enumerate(low) if "awb" in c), None)
                ticket_i = next((i for i, c in enumerate(low) if "ticket" in c), None)
                status_i = si
                active = True
            else:
                active = False                                # quoted original
            continue
        if not active:
            continue
        need = max(i for i in (awb_i, ticket_i, status_i) if i is not None)
        if len(cells) <= need:
            continue
        status = cells[status_i].strip()
        awb = norm_awb(cells[awb_i]) if awb_i is not None else ""
        ticket = cells[ticket_i].strip() if ticket_i is not None else ""
        if not status:
            continue
        if not re.fullmatch(r"\d{4,}", ticket):
            ticket = ""
        if not re.fullmatch(r"\d{6,}", awb):
            awb = ""
        if ticket or awb:
            rows.append({"ticket": ticket, "awb": awb, "status": status})
    return rows


def html_of(msg):
    for part in msg.walk():
        if part.get_content_type() == "text/html":
            try:
                return part.get_payload(decode=True).decode(
                    part.get_content_charset() or "utf-8", "ignore")
            except Exception:
                continue
    return ""


def fetch_message(C, uid):
    typ, d = C.uid("FETCH", uid, "(RFC822)")
    for item in d or []:
        if isinstance(item, tuple) and len(item) > 1 and isinstance(item[1], (bytes, bytearray)):
            return email.message_from_bytes(item[1])
    return None


def header_date(raw):
    """A Date: header as an aware datetime, or None when it cannot be read.

    A naive value is read as IST rather than thrown away: a missing offset is far more
    often a client that omitted one than a message from another zone, and an aware
    value is required for the comparison below to work at all."""
    if not raw:
        return None
    try:
        d = parsedate_to_datetime(raw)
    except (TypeError, ValueError):
        return None
    if d is None:
        return None
    return d if d.tzinfo else d.replace(tzinfo=IST)


def find_thread_replies(C, thread, siblings=()):
    """UIDs of replies to THIS thread, oldest first.

    Matching is done CLIENT-SIDE on the fetched headers. Zoho's IMAP answers
    `SEARCH HEADER "In-Reply-To" ...` with OK and zero hits even when the message is
    plainly there, so a server-side header search silently finds nothing. Only the
    date range is pushed to the server; the decision stays here.

    A reply belongs to exactly ONE report, and these rules decide which, in order:

    1. It quotes our Message-ID in In-Reply-To/References -> it is that report's,
       full stop. This is the authoritative case and covers every normal mail client.

    2. It quotes a DIFFERENT report of ours -> it belongs to that one, so this thread
       declines it. Without this rule the subject fallback below would claim it too:
       every report shares the fixed subject prefix, and reports sent on the same day
       share the subject byte for byte, so an answer to the morning report would also
       be applied to the afternoon one and mark it answered although the courier
       never saw it.

    3. It quotes none of ours and merely repeats the subject (a client that stripped
       the headers) -> it is booked to the NEAREST REPORT THAT PRECEDES IT, since
       that is the one it was most likely written against. Every other same-subject
       thread declines it, so it is applied once and only once. A reply older than a
       report can never be that report's. If its date is unreadable it is accepted
       only when this is the sole thread carrying that subject; otherwise there is
       nothing left to tell them apart and it is left alone rather than guessed at."""
    C.select("INBOX", readonly=True)
    sent = datetime.fromisoformat(thread["sent_ist"])
    since = (sent - timedelta(days=1)).strftime("%d-%b-%Y")
    try:
        typ, data = C.uid("SEARCH", None, '(SINCE "{}")'.format(since))
        if typ != "OK":
            return []
    except Exception:
        return []

    mid = thread["message_id"].strip()
    subject_tail = thread["subject"].strip().lower()
    # Every other report we have sent: their ids claim a reply away from this thread,
    # and those sharing this subject compete for a header-less one.
    others = [t for t in siblings if t is not thread]
    other_mids = {(t.get("message_id") or "").strip() for t in others} - {"", mid}
    same_subject = [t for t in others
                    if (t.get("subject") or "").strip().lower() == subject_tail]

    out = []
    for uid in (data[0] or b"").split():
        typ, d = C.uid("FETCH", uid,
                       "(BODY.PEEK[HEADER.FIELDS (MESSAGE-ID IN-REPLY-TO REFERENCES SUBJECT DATE)])")
        item = next((i for i in (d or []) if isinstance(i, tuple)), None)
        if not item:
            continue
        h = email.message_from_bytes(item[1])
        if (h.get("Message-ID") or "").strip() == mid:
            continue                                   # our own sent copy
        refs = "{} {}".format(h.get("In-Reply-To") or "", h.get("References") or "")
        if mid in refs:                                 # rule 1
            out.append(uid)
            continue
        if any(o in refs for o in other_mids):          # rule 2 - another report's
            continue
        subj = str(make_header(decode_header(h.get("Subject") or ""))).strip().lower()
        if subj.endswith(subject_tail) and subj != subject_tail:
            when = header_date(h.get("Date"))           # rule 3
            if when is None:
                if same_subject:
                    continue                            # cannot be told apart
            else:
                if when < sent:
                    continue                            # predates this report
                nearer = [t for t in same_subject
                          if sent < datetime.fromisoformat(t["sent_ist"]) <= when]
                if nearer:
                    continue                            # a later report fits better
            msg = fetch_message(C, uid)
            if msg is not None and parse_status_table(html_of(msg)):
                out.append(uid)
    return sorted(out, key=lambda u: int(u))


def resolve_ticket_id(ticket_number, max_pages=20):
    """Internal ticket id from a ticket number, by scanning the ticket list.

    The search API needs a scope this token does not have (403), so this pages
    newest-first until it finds the number. Only used for AWBs registered before ids
    were captured; the result is written back to the registry so it happens once."""
    frm = 1
    for _ in range(max_pages):
        d = api_get("tickets?limit=100&from={}&sortBy=-createdTime".format(frm))
        rows = d.get("data") or []
        if not rows:
            return None
        for t in rows:
            if str(t.get("ticketNumber")) == str(ticket_number):
                return t.get("id")
        frm += 100
    return None


def thread_targets(thread):
    """{ticketNumber: {"ticketId", "awb"}} for a thread, new or old record shape."""
    if thread.get("tickets"):
        return thread["tickets"]
    out = {}
    for awb, v in (thread.get("awbs") or {}).items():
        num = str(v.get("ticketNumber") or "").strip()
        if num:
            out[num] = {"ticketId": v.get("ticketId"), "awb": awb}
    return out


def pending_targets(thread):
    """{ticketNumber: meta} for the lines on this thread still worth chasing.

    A line drops out only once its remark was FINAL. Being answered is not the
    same as being finished - "OFD" is an answer and still needs chasing, which is
    the entire point of this cycle. Keeping the two apart is why finality is
    tracked in its own dict rather than reusing `processed`."""
    final = thread.get("final") or {}
    return {num: meta for num, meta in thread_targets(thread).items()
            if num not in final}


def locate_target(threads, current, row, registry):
    """Resolve one reply row to the ticket it belongs to.

    THE TICKET NUMBER IS THE PREFERRED KEY. It is unique, and we put it in the mail
    precisely so that it comes back to us, so a row carrying one is resolved on that
    alone and the AWB is never consulted.

    The AWB is the fallback for the rows that have no ticket number, which is most of
    them in practice: Bluedart usually answer on their own template, which has an AWB
    column and no ticket column. Requiring a ticket number there dropped every row of
    every such reply, so nothing was ever commented. An AWB is NOT unique - one
    shipment can carry several tickets - so it only decides when it resolves to
    exactly one ticket among the threads awaiting a reply; anything ambiguous is
    reported and skipped, never guessed at.

    Bluedart also answers on whichever mail is nearest to hand, so a reply to Monday's
    report routinely carries rows from Wednesday's. Every open thread is searched, and
    the thread that actually owns the ticket is returned so the update is booked
    against it rather than against whichever thread we happened to be reading.

    Returns (owner_thread_or_None, {"ticketNumber","ticketId"} or None, description)."""
    order = [current] + [t for t in threads if t is not current]

    def where(t):
        return "this thread" if t is current else "thread {}".format(t.get("subject"))

    num = (row.get("ticket") or "").strip()
    if not num:
        # Bluedart answer on their own template: AWB and a status, no ticket column.
        # An AWB is NOT unique - the same shipment can carry several tickets - so it
        # is trusted ONLY when it resolves to exactly one ticket across the threads we
        # are still waiting on. More than one hit is reported and skipped rather than
        # guessed at, which is the case the ticket-number rule was protecting against.
        awb = norm_awb(row.get("awb") or "")
        if not awb:
            return None, None, "no ticket number and no AWB in the reply row"
        hits = {}
        for t in order:
            for n, meta in thread_targets(t).items():
                if norm_awb(meta.get("awb") or "") == awb and n not in hits:
                    hits[n] = (t, meta)
        if len(hits) == 1:
            n, (t, meta) = next(iter(hits.items()))
            return t, {"ticketNumber": n, "ticketId": meta.get("ticketId")}, \
                "AWB {} -> #{} on {}".format(awb, n, where(t))
        if len(hits) > 1:
            return None, None, ("AWB {} matches {} tickets ({}) - ambiguous, skipped"
                                .format(awb, len(hits), ", ".join(sorted(hits))))
        return None, None, "AWB {} is not on any thread awaiting a reply".format(awb)

    for t in order:
        meta = thread_targets(t).get(num)
        if meta:
            return t, {"ticketNumber": num, "ticketId": meta.get("ticketId")}, where(t)

    # The registry is keyed by AWB, but this is still a ticket-number lookup: it scans
    # the entries for that number. The AWB is not what selects the ticket.
    for awb, entry in registry.items():
        if str(entry.get("ticketNumber") or "") == num:
            tid = entry.get("ticketId") or resolve_ticket_id(num)
            if tid:
                if not entry.get("ticketId"):
                    entry["ticketId"] = tid
                    save_registry(registry)
                return None, {"ticketNumber": num, "ticketId": tid}, "registry (by ticket no.)"

    tid = resolve_ticket_id(num)
    if tid:
        return None, {"ticketNumber": num, "ticketId": tid}, "Desk lookup"
    return None, None, "unknown"


def comment_on_ticket(ticket_id, ticket_number, status, when):
    """Private comment. isPublic false IS the Private button.

    Addressed AND labelled by the ticket alone. The AWB is deliberately kept out of
    both: it is not unique, so naming it here would imply the update had been matched
    on it, when the ticket number is what actually selected this ticket."""
    content = ("Bluedart update ({date})\nTicket #{num}: {status}\n\n"
               "Added automatically from the escalation email reply."
               ).format(date=when, num=ticket_number, status=status)
    return api_post(COMMENT_PATH.format(ticket_id),
                    {"content": content, "isPublic": "false", "contentType": "plainText"})


def followup_body(thread):
    """(text, html) for a chase naming only the AWBs still not delivered.

    The table is the SAME shape we escalated with - every mapping column, in the
    same order - so the trail reads as one conversation and Bluedart can answer
    the chase exactly as they answered the first mail. Only the rows that reached
    a delivered status are dropped.

    Any status-like column is stripped from what we send, however it reached the
    mapping. parse_status_table reads "our own headers plus one extra column" as
    a filled-in reply, so echoing a status column back would make our own chase
    parse as a Bluedart status the moment they quote it - posting a comment,
    re-arming the timer off our own words, and chasing forever. The status column
    is theirs to add, never ours to send.

    Threads recorded before the mapping row was persisted have no `columns`; they
    fall back to ticket number and AWB, which is all such a thread ever stored."""
    pending = pending_targets(thread)
    book = thread.get("processed") or {}
    order = sorted(pending)

    cols = list(thread.get("columns") or [])
    grid = [list(pending[num].get("row") or []) for num in order]

    if cols and grid and all(len(g) == len(cols) for g in grid):
        drop = _status_index(cols)
        if drop is not None:
            cols = [c for i, c in enumerate(cols) if i != drop]
            grid = [[v for i, v in enumerate(g) if i != drop] for g in grid]
    else:
        cols = ["Ticket Number", "AWB Number"]
        grid = [[num, pending[num].get("awb") or "-"] for num in order]

    text = "{}\n\n{}\n".format(FOLLOWUP_TEXT, "\n".join(
        ["  |  ".join(cols)] +
        ["  |  ".join("" if v is None else str(v) for v in g) for g in grid]))

    # The remark they last gave goes in prose, not in a column: a cell would
    # rebuild the very shape that made our own mail parse as a reply.
    echoed = ["#{} - {}".format(_esc(num), _esc((book.get(num) or {}).get("status")))
              for num in order if (book.get(num) or {}).get("status")]
    html = (
        '<div style="font-family:Segoe UI,Arial,sans-serif;font-size:14px;'
        'color:#000000;"><p>{}</p>'
        '<table cellpadding="6" cellspacing="0" border="1" '
        'style="border-collapse:collapse;font-size:13px;">'
        '<tr style="background:#f2f2f2;">{}</tr>{}</table>{}</div>'
    ).format(
        _esc(FOLLOWUP_TEXT),
        "".join("<th>{}</th>".format(_esc(c)) for c in cols),
        "".join("<tr>{}</tr>".format(
            "".join("<td>{}</td>".format(_esc("" if v is None else str(v)))
                    for v in g)) for g in grid),
        "<p>Your last update: {}.</p>".format("; ".join(echoed)) if echoed else "")
    return text, html


def addr_list(key):
    """A comma-separated .env address list as a list. Persistence only."""
    return [a.strip() for a in (ENV.get(key) or "").split(",") if a.strip()]


def send_followup(thread):
    """Chase on the same thread, so it lands in the existing conversation."""
    import smtplib
    from email.message import EmailMessage

    sender = (ENV.get("MAIL_FROM") or "").strip() or ENV["OTP_EMAIL"]
    to = [a.strip() for a in ENV["MAIL_TO"].split(",") if a.strip()]
    cc = [a.strip() for a in (ENV.get("MAIL_CC") or "").split(",") if a.strip()]

    msg = EmailMessage()
    msg["From"] = sender
    msg["To"] = ", ".join(to)
    if cc:
        msg["Cc"] = ", ".join(cc)
    msg["Subject"] = "Re: " + thread["subject"]
    msg["In-Reply-To"] = thread["message_id"]
    msg["References"] = thread["message_id"]
    text, html = followup_body(thread)
    msg.set_content(text)
    msg.add_alternative(html, subtype="html")
    try:
        with smtplib.SMTP(ENV["SMTP_HOST"], int(ENV["SMTP_PORT"]), timeout=60) as s:
            s.starttls()
            s.login(ENV["OTP_EMAIL"], ENV["OTP_EMAIL_PASSWORD"])
            s.send_message(msg, from_addr=sender, to_addrs=to + cc)
        print("  follow-up sent on thread {}".format(thread["subject"]))
        return True
    except Exception as e:
        print("  follow-up FAILED: {}: {}".format(e.__class__.__name__, e))
        return False


def reply_received_at(msg):
    """When Bluedart's reply landed in our mailbox.

    The Date header is what the requirement means by "received". The 15-hour clock
    hangs off it rather than off when this run happened to read the mail, so a run
    delayed by an outage does not stretch the interval to match."""
    try:
        d = parsedate_to_datetime(msg.get("Date"))
        if d is not None:
            if d.tzinfo is None:
                d = d.replace(tzinfo=IST)
            return d.astimezone(IST)
    except Exception:
        pass
    return datetime.now(IST)


def followup_interval_hours():
    """Hours to wait after a reply before chasing again.

    REPLY_FOLLOWUP_HOURS is the production cadence and the default. A test run
    sets REPLY_FOLLOWUP_HOURS=0 in .env to chase at once rather than sit out the
    real interval. Anything unparseable falls back to the default rather than
    raising, so a fat-fingered .env cannot stop the pipeline chasing; a negative
    is clamped to 0 so a chase is never armed into the past."""
    raw = str((ENV or {}).get("REPLY_FOLLOWUP_HOURS") or "").strip()
    if not raw:
        return REPLY_FOLLOWUP_HOURS
    try:
        return max(0, float(raw))
    except ValueError:
        print("  warning: REPLY_FOLLOWUP_HOURS={!r} is not a number - "
              "using {}h".format(raw, REPLY_FOLLOWUP_HOURS))
        return REPLY_FOLLOWUP_HOURS


def db_email_id(thread):
    """emails_sent.id for this thread, or None when persistence is off.

    Cached on the thread so the chase path does not re-query on every tick."""
    got = thread.get("db_email_id")
    if got:
        return got
    got = db.email_id_for(thread.get("message_id"))
    if got:
        thread["db_email_id"] = got
    return got


def pause_followups(thread, received):
    """Bluedart answered, but not with a status. Stand the chase down.

    Chasing someone who has just written to us reads as though we ignored them,
    and their next mail is usually the real status. So a reply carrying no status
    table PAUSES the cycle rather than leaving the old timer to fire: nothing more
    is sent until a reply that actually carries a table re-arms it.

    The 15:00 no-reply net stands down too, because last_reply_ist is now set -
    Bluedart has replied, whatever the content."""
    thread["last_reply_ist"] = received.isoformat()
    thread["next_followup_ist"] = None
    thread["awaiting_status_since"] = received.isoformat()
    db.set_followup_state(db_email_id(thread), last_reply_at=received,
                          next_followup_at=None,
                          awaiting_status_since=received)


def arm_followup(thread, received):
    """Point the next chase at 15 hours after this reply landed.

    Disarmed once nothing is pending or the cap is hit, so a finished or exhausted
    thread cannot be woken up again by reply_followup_due."""
    thread["last_reply_ist"] = received.isoformat()
    thread["awaiting_status_since"] = None      # a real status lifts the pause
    if (pending_targets(thread)
            and int(thread.get("followup_round") or 0) < FOLLOWUP_MAX_ROUNDS):
        thread["next_followup_ist"] = (
            received + timedelta(hours=followup_interval_hours())).isoformat()
    else:
        thread["next_followup_ist"] = None
    db.set_followup_state(
        db_email_id(thread), last_reply_at=received,
        awaiting_status_since=None,
        next_followup_at=(datetime.fromisoformat(thread["next_followup_ist"])
                          if thread["next_followup_ist"] else None))


def reply_followup_due(thread, now=None):
    """True when the 15-hour reply-anchored chase is ready to go out.

    Armed only by an actual reply landing, so a thread Bluedart has never answered
    is left to the 15:00 net below instead of being chased by both mechanisms."""
    if thread.get("followup_suppressed"):
        return False
    nxt = thread.get("next_followup_ist")
    if not nxt:
        return False
    if int(thread.get("followup_round") or 0) >= FOLLOWUP_MAX_ROUNDS:
        return False
    if not pending_targets(thread):
        return False
    return (now or datetime.now(IST)) >= datetime.fromisoformat(nxt)


def followup_due(thread, now=None):
    """True once it is past FOLLOWUP_HOUR on the day AFTER the report went out.

    The no-reply safety net, and nothing more. Once Bluedart has answered even
    once, the 15-hour cycle owns the thread and this rule steps aside - otherwise
    a thread mid-cycle would be chased by both."""
    if thread.get("followup_suppressed") or thread.get("last_reply_ist"):
        return False
    if thread.get("status") != "awaiting_reply" or thread.get("followup_sent_ist"):
        return False
    now = now or datetime.now(IST)
    sent = datetime.fromisoformat(thread["sent_ist"])
    due = (sent + timedelta(days=1)).replace(hour=FOLLOWUP_HOUR, minute=0,
                                             second=0, microsecond=0)
    return now >= due


def push_to_analytics():
    """Send any ticket whose journey moved to Zoho Analytics. Best-effort.

    Runs as a SEPARATE PROCESS on purpose. The escalation pipeline must never
    fail because a reporting push did - an expired token or a Zoho outage costs
    a report, an escalation that does not go out costs a customer - and a child
    process cannot take this one down with it.

    Silent no-op when Analytics is not configured, so a checkout without ZA_*
    keys behaves exactly as it did before."""
    script = BASE / "db" / "push_analytics.py"
    if not script.exists() or not (ENV.get("ZA_REFRESH_TOKEN") or "").strip():
        return
    try:
        r = subprocess.run([sys.executable, str(script)], cwd=str(BASE),
                           capture_output=True, text=True, timeout=180)
        out = (r.stdout or "").strip()
        if out:
            print(out)
        if r.returncode and (r.stderr or "").strip():
            print("  analytics: {}".format(r.stderr.strip().splitlines()[-1][:160]))
    except Exception as e:
        # Including a timeout. Never re-raised: see the docstring.
        print("  analytics push skipped ({}: {})".format(e.__class__.__name__, e))


def process_replies(verbose=True, dry=False):
    """Apply any new reply statuses to their tickets. Returns count of comments made.

    Safe to run repeatedly: a thread records which AWBs it has already commented, so
    re-reading the same reply adds nothing.

    dry=True resolves everything and prints what WOULD be written, touching neither a
    ticket nor the thread file. Since it changes nothing it also inspects threads that
    are already completed, so a finished mapping can still be reviewed."""
    # A dry run writes nothing, so it needs no lock and must not block a real one.
    if dry:
        return _process_replies(verbose=verbose, dry=True)
    # Held across the WHOLE scan - load through save - because that span is exactly
    # the window in which another process's append would be lost. webhook_server.py
    # drives this from a separate process, and --watch from another again.
    with threads_lock(label="process_replies"):
        made = _process_replies(verbose=verbose, dry=False)
    # Outside the lock deliberately: the push reads Postgres, not threads.json,
    # and holding the file lock across an HTTP call would stall the daily run.
    push_to_analytics()
    return made


def _process_replies(verbose=True, dry=False):
    global TOKEN
    threads = load_threads()
    open_threads = [t for t in threads
                    if dry or t.get("status") == "awaiting_reply"]
    if not open_threads:
        if verbose:
            print("No open threads.")
        return 0

    if TOKEN is None:
        TOKEN = get_token()

    registry = load_registry()          # global AWB -> ticket, the last-resort lookup
    C = imap_connect(quiet=not verbose)
    made = 0
    try:
        for thread in open_threads:
            # Every thread is passed in, not just the open ones: a reply that names a
            # report we have already closed belongs to that one, and must not be
            # re-applied to a newer report that happens to share its subject.
            uids = find_thread_replies(C, thread, threads)
            if verbose:
                print("\nThread {} - {} reply/replies".format(
                    thread["subject"], len(uids)))
            thread.setdefault("processed", {})
            seen = thread.setdefault("seen_replies", [])

            for uid in uids:
                msg = fetch_message(C, uid)
                if msg is None:
                    continue
                rid = (msg.get("Message-ID") or "uid:{}".format(uid.decode())).strip()
                if rid in seen and not dry:
                    continue          # a dry run re-reads them, since it writes nothing
                statuses = parse_status_table(html_of(msg))
                if verbose:
                    print("  reply from {} - {} status row(s)".format(
                        msg.get("From"), len(statuses)))
                received = reply_received_at(msg)
                if not statuses:
                    # Not nothing: they answered. Pause and wait for the status
                    # rather than letting the previous timer chase them anyway.
                    if verbose:
                        print("    no Status table - they have replied but not with "
                              "a status; chase paused until they send one")
                    if not dry:
                        pause_followups(thread, received)
                        seen.append(rid)
                    continue

                when = received.strftime(MAIL_DATE_FMT)
                # The reply, body and all, before any row is handled. message_id
                # is the idempotency key, so re-reading one inserts nothing.
                reply_id = None if dry else db.record_reply(
                    email_id=db_email_id(thread), message_id=rid,
                    in_reply_to=(msg.get("In-Reply-To") or "").strip() or None,
                    from_addr=str(msg.get("From") or "") or None,
                    subject=str(msg.get("Subject") or "") or None,
                    received_at=received, body_html=html_of(msg),
                    status_rows_found=len(statuses))
                touched = []
                done_this_reply = set()
                for row in statuses:
                    awb, status = row.get("awb", ""), row["status"]
                    owner, target, via = locate_target(threads, thread, row, registry)
                    if not target or not target.get("ticketId"):
                        print("    SKIP ticket {!r} / AWB {!r} - {}".format(
                            row.get("ticket"), awb, via))
                        continue
                    key = str(target["ticketNumber"])
                    # Keyed on the ticket number, not the AWB: one AWB can have
                    # several tickets, so an AWB key would collapse them into one.
                    # Booked against the thread that owns the ticket, so that
                    # report closes properly.
                    # One comment per ticket per REPLY. Bluedart sometimes lists a
                    # ticket twice in one table; without this the row would be
                    # commented once per occurrence.
                    if key in done_this_reply:
                        print("    #{} appears twice in this reply - keeping the "
                              "first remark".format(key))
                        continue
                    done_this_reply.add(key)
                    home = owner or thread
                    book = home.setdefault("processed", {})
                    final = home.setdefault("final", {})
                    # EVERY remark is commented, including the second and third on a
                    # ticket already answered: "OFD" -> "Under follow up" ->
                    # "Delivered" is the normal progression and the ticket has to
                    # carry all of it. Re-reading the same reply is prevented by
                    # seen_replies above, which is the right dedup key - one per
                    # reply, not one per ticket for all time.
                    if owner is not thread:
                        print("    note: ticket #{} belongs to {}".format(key, via))
                    if dry:
                        print("    WOULD COMMENT -> ticket #{}  (id {})".format(
                            key, target.get("ticketId")))
                        print("      AWB {}   remark: {!r}   via {}".format(
                            awb or "-", status, via))
                        made += 1
                        continue
                    try:
                        posted = comment_on_ticket(target["ticketId"], key,
                                                   status, when)
                    except Exception as e:
                        print("    FAILED ticket {}: {}".format(key, e))
                        continue
                    # comment_on_ticket's return was previously discarded. Desk
                    # answers with the comment it created, so this is the only
                    # place its id can be captured for the record.
                    comment_id = (posted or {}).get("id") \
                        if isinstance(posted, dict) else None
                    book[key] = {"status": status, "awb": awb, "via": via,
                                 "commented_ist": datetime.now(IST).isoformat()}
                    # Finality is recorded once and never lifted: a later non-final
                    # remark on a delivered AWB is still commented, but must not
                    # drag it back into the chase.
                    if key not in final and is_final_status(status):
                        final[key] = {"status": status, "awb": awb,
                                      "finalised_ist": received.isoformat()}
                        print("    #{} is final ({!r}) - no further chasing".format(
                            key, status))
                    # Append-only history, then the cached latest on the line.
                    # These are the two grains the JSON store could not keep:
                    # it overwrote the remark and lost everything before it.
                    db.record_status(reply_id, key, status, awb=awb,
                                     matched_via=via, received_at=received,
                                     desk_comment_id=comment_id,
                                     desk_posted_at=datetime.now(IST))
                    db.apply_latest(db_email_id(home), key, status,
                                    reply_id=reply_id,
                                    desk_comment_id=comment_id,
                                    desk_posted_at=datetime.now(IST),
                                    is_final=is_final_status(status),
                                    final_reason=(
                                        None if not is_final_status(status)
                                        else "delivered"
                                        if is_delivered_status(status)
                                        else "rto"))
                    # The two outcomes are stamped separately. Narrower than
                    # finality on purpose: plain RTO is final but nothing was
                    # delivered, while "RTO Delivered" was. Both carry the
                    # reply's own received time, and both keep the first one
                    # forever.
                    if is_delivered_status(status):
                        db.mark_delivered(key, received, status_text=status,
                                          awb=awb, reply_id=reply_id)
                    elif is_final_status(status):
                        db.mark_rto(key, received, status_text=status,
                                    awb=awb, reply_id=reply_id)
                    if home not in touched:
                        touched.append(home)
                    made += 1
                    print("    #{} <- {!r}  (AWB {})".format(key, status, awb or "-"))
                if not dry:
                    for home in touched:
                        arm_followup(home, received)
                seen.append(rid)

        # Completion is judged across ALL threads, not just the one being read: a
        # reply on Monday's mail can finish Wednesday's report by answering its last
        # outstanding AWB.
        for t in threads:
            if t.get("status") != "awaiting_reply":
                continue
            answered = t.get("processed") or {}
            # Outstanding means NOT DELIVERED, not merely "not yet answered". An AWB
            # answered "OFD" is answered and still outstanding - closing the thread
            # on it is exactly the bug this cycle exists to fix.
            outstanding = sorted(pending_targets(t))
            if answered and not dry:
                # Derived in SQL from the lines, so the database agrees with the
                # thread file rather than being told separately.
                db.rollup_email(db_email_id(t))
            if answered and not outstanding:
                if not dry:
                    t["status"] = "completed"
                    t["completed_ist"] = datetime.now(IST).isoformat()
                    t["next_followup_ist"] = None
                print("  thread completed ({}): all {} AWB(s) delivered".format(
                    t.get("subject"), len(thread_targets(t))))
            elif answered:
                print("  thread still open ({}): {} AWB(s) not delivered: {}".format(
                    t.get("subject"), len(outstanding), ", ".join(outstanding)))
    finally:
        try:
            C.logout()
        except Exception:
            pass
        if not dry:
            save_threads(threads)
    return made


def run_followups(verbose=True, now=None):
    """Send whatever chases are due, by either mechanism.

    Two independent rules, and a thread is only ever eligible for one of them:

      * the 15-hour reply-anchored cycle, for a thread Bluedart HAS answered but
        which still has AWBs short of a delivered status;
      * the 15:00 next-day nudge, for a thread Bluedart has NEVER answered.

    The cycle re-arms itself on send rather than only on reply, so a thread that
    goes quiet after one answer keeps being chased instead of stalling with AWBs
    still open. At the cap it disarms and is left alone."""
    # Locked: without it a concurrent reply scan can write back a copy loaded
    # before followup_sent_ist was stamped, and the courier is nudged twice.
    with threads_lock(label="run_followups"):
        sent = _run_followups(verbose, now)
    push_to_analytics()
    return sent


def _run_followups(verbose=True, now=None):
    threads = load_threads()
    now = now or datetime.now(IST)
    sent = 0
    for thread in threads:
        if reply_followup_due(thread, now):
            # Captured BEFORE the send: these are what this chase went out with.
            # followup_body is the same pure call send_followup makes, from the
            # same thread state, so it cannot fail here without failing there.
            chased = sorted(pending_targets(thread))
            body_text, body_html = followup_body(thread)
            if not send_followup(thread):
                continue                  # left armed in the past; retried next tick
            # Appended before the round is incremented, so the row describes the
            # chase that actually went out rather than the next one.
            db.record_followup(
                db_email_id(thread),
                round=int(thread.get("followup_round") or 0) + 1,
                kind="reply_anchored", sent_at=now,
                to_addrs=addr_list("MAIL_TO"), cc_addrs=addr_list("MAIL_CC"),
                subject="Re: " + str(thread.get("subject") or ""),
                body_text=body_text, body_html=body_html,
                ticket_numbers=chased)
            thread["followup_round"] = int(thread.get("followup_round") or 0) + 1
            thread["followup_sent_ist"] = now.isoformat()
            if thread["followup_round"] >= FOLLOWUP_MAX_ROUNDS:
                thread["next_followup_ist"] = None
                thread["stalled_ist"] = now.isoformat()
                db.set_followup_state(db_email_id(thread), stalled_at=now,
                                      next_followup_at=None,
                                      followup_round=thread["followup_round"],
                                      followup_sent_at=now)
                print("  {} chases with no delivered status - leaving {} alone: "
                      "{}".format(FOLLOWUP_MAX_ROUNDS, thread.get("subject"),
                                  ", ".join(sorted(pending_targets(thread)))))
            else:
                thread["next_followup_ist"] = (
                    now + timedelta(hours=followup_interval_hours())).isoformat()
                db.set_followup_state(
                    db_email_id(thread),
                    next_followup_at=datetime.fromisoformat(
                        thread["next_followup_ist"]),
                    followup_round=thread["followup_round"],
                    followup_sent_at=now)
            sent += 1
        elif followup_due(thread, now):
            chased = sorted(pending_targets(thread))
            body_text, body_html = followup_body(thread)
            if send_followup(thread):
                thread["followup_sent_ist"] = now.isoformat()
                db.set_followup_state(db_email_id(thread), followup_sent_at=now)
                db.record_followup(
                    db_email_id(thread),
                    round=int(thread.get("followup_round") or 0),
                    kind="no_reply", sent_at=now,
                    to_addrs=addr_list("MAIL_TO"),
                    cc_addrs=addr_list("MAIL_CC"),
                    subject="Re: " + str(thread.get("subject") or ""),
                    body_text=body_text, body_html=body_html,
                    ticket_numbers=chased)
                sent += 1
    if sent:
        save_threads(threads)
    elif verbose:
        print("No follow-ups due.")
    return sent


def watch(poll_fallback=60):
    """Stay connected and react the moment a reply lands.

    IMAP IDLE is a server push, so a reply is picked up in a second or two without a
    public webhook endpoint. The IDLE window is capped below the 29-minute protocol
    limit; each wake also re-checks whether a follow-up has come due."""
    print("Watching for replies. Ctrl+C to stop.")
    print("  never answered: one nudge at {:02d}:00 IST the day after a report"
          .format(FOLLOWUP_HOUR))
    print("  answered, still not delivered: chased every {}h on the same "
          "trail, up to {} times".format(REPLY_FOLLOWUP_HOURS,
                                         FOLLOWUP_MAX_ROUNDS))
    while True:
        try:
            process_replies(verbose=True)
            run_followups(verbose=False)

            C = imap_connect()
            try:
                C.select("INBOX")
                if hasattr(C, "idle"):
                    print("[{:%H:%M:%S}] idle...".format(datetime.now(IST)))
                    with C.idle(duration=15 * 60) as idler:
                        for typ, data in idler:
                            if b"EXISTS" in bytes(str(data), "utf-8") or typ == "EXISTS":
                                print("  new mail signalled")
                                break
                else:
                    # imaplib only grew IMAP4.idle() in Python 3.14. On an older
                    # interpreter there is no push channel at all, so poll instead:
                    # replies land a little later, nothing else about the loop changes.
                    print("[{:%H:%M:%S}] polling every {}s "
                          "(Python {}.{} has no IMAP IDLE)".format(
                              datetime.now(IST), poll_fallback,
                              sys.version_info[0], sys.version_info[1]))
                    time.sleep(poll_fallback)
            finally:
                try:
                    C.logout()
                except Exception:
                    pass
        except KeyboardInterrupt:
            print("\nStopped.")
            return 0
        except Exception as e:
            print("watcher hiccup: {}: {} - retrying in {}s".format(
                e.__class__.__name__, e, poll_fallback))
            time.sleep(poll_fallback)


# ============================================================================
# Self-test for the OTP extractor (no browser, no network)
# ============================================================================

REAL_OTP_MAIL = (
    "Your ClickPost Login OTP Login OTP Verification Hello account.holder, "
    "Your One-Time Password (OTP) for ClickPost login is: X63YUN This OTP is valid "
    "for 10 minutes only. If you did not request this OTP, please ignore this email.")
REAL_REPORT_MAIL = (
    "ClickPost | Untitled Report | Aug 19, 2026 96 Hi, Here is the report that you "
    "requested. CLICK POST | Reports Untitled Report Report ID: 361870 Generated 2026")
HTML_OTP_MAIL = ("<html><head><style>.x{color:#fff}</style></head><body><p>Hello,</p>"
                 "<p>Your One-Time Password (OTP) is: <b>729184</b></p>"
                 "<span>Valid till Aug 19, 2026 10:54</span></body></html>")

SELFTESTS = [
    ("real ClickPost OTP mail (alphanumeric)", REAL_OTP_MAIL, "X63YUN"),
    ("real ClickPost OTP mail, second code",
     REAL_OTP_MAIL.replace("X63YUN", "W8BV3V"), "W8BV3V"),
    ("report mail yields nothing", REAL_REPORT_MAIL, None),
    # 2026-08-24: a real code with no digits at all, which the old digit
    # requirement silently rejected - the login then timed out with the mail
    # sitting unread in the mailbox.
    ("all-letter code (no digits)",
     REAL_OTP_MAIL.replace("X63YUN", "KWUUHG"), "KWUUHG"),
    ("all-letter code must not win without OTP context",
     "Quarterly summary attached. Totals are FINAL and approved.", None),
    ("numeric OTP still works",
     "Your OTP for ClickPost login is: 483920. Valid for 10 minutes.", "483920"),
    ("verification code wording",
     "Use this verification code to sign in: 274815", "274815"),
    ("one-time password, hyphenated",
     "Hi, your one-time password is 55213 - do not share it.", "55213"),
    ("year never wins over the real code",
     "ClickPost OTP | Aug 19, 2026. Your one-time password is: 601233", "601233"),
    ("no OTP context at all", "Your ClickPost report for Aug 19, 2026 is ready.", None),
    ("empty body", "", None),
    ("bare year rejected", "OTP mail 2026", None),
    # A code may be any mix of letters and digits. These cover the shapes that the
    # old digit requirement rejected, including the unanchored ones.
    ("letters-only, labelled with a colon",
     "ClickPost security code. Your OTP: MJXQVB", "MJXQVB"),
    ("letters-only, no anchoring label at all",
     "Your one-time password for ClickPost login, quoting reference MJXQVB below.",
     "MJXQVB"),
    ("letters-only wins over a capitalised English word",
     "Your OTP was issued. Please LOGIN using MJXQVB now.", "MJXQVB"),
    ("a token with a digit is preferred when both are present",
     "Your OTP details. Reference ABCDEF, code 4KJ2QP.", "4KJ2QP"),
    ("digits-only still works unanchored",
     "Your one-time password, reference 481920 for this login.", "481920"),
]


def run_selftest():
    failed = 0
    for desc, body, expected in SELFTESTS:
        got = extract_otp(body)
        ok = got == expected
        failed += not ok
        print("  [{}] {}: got {!r}, expected {!r}".format(
            "PASS" if ok else "FAIL", desc, got, expected))
    got = extract_otp(html_to_text(HTML_OTP_MAIL))
    ok = got == "729184"
    failed += not ok
    print("  [{}] html body: got {!r}, expected '729184'".format(
        "PASS" if ok else "FAIL", got))
    print("\n{}".format("ALL PASSED" if not failed else "{} FAILED".format(failed)))
    return failed


# ============================================================================
# Pipeline
# ============================================================================


def flag(name):
    return name in sys.argv[1:]


def opt(name, default=None):
    argv = sys.argv[1:]
    if name in argv and argv.index(name) + 1 < len(argv):
        return argv[argv.index(name) + 1]
    return default


def awb_count(xlsx):
    from openpyxl import load_workbook
    ws = load_workbook(xlsx)[SRC_SHEET]
    header = [c.value for c in ws[1]]
    idx = header.index(AWB_TICKETS)
    return sum(1 for r in ws.iter_rows(min_row=2) if r[idx].value)


def main():
    global ENV
    if flag("--selftest"):
        return 1 if run_selftest() else 0

    # Reply-side modes. These never touch Desk tickets other than the ones recorded
    # on our own threads, and never read mail outside those threads.
    if flag("--watch") or flag("--process-replies") or flag("--followups"):
        ENV = load_env()
        if flag("--watch"):
            return watch()
        rc = 0
        if flag("--process-replies"):
            n = process_replies(dry=flag("--dry"))
            print("\n{} comment(s) added.".format(n))
        if flag("--followups"):
            n = run_followups()
            print("{} follow-up(s) sent.".format(n))
        return rc

    started = time.time()
    OUT.mkdir(parents=True, exist_ok=True)
    ENV = load_env()
    state = load_state()
    watermark = None            # set when this run owns an uncommitted watermark

    if flag("--abandon-pending"):
        p = state.pop("pending", None)
        save_state(state)
        print("Abandoned pending run: {}".format(
            p.get("export_file") if p else "there was none"))
        return 0

    # ---- Part 1: Zoho Desk -------------------------------------------------
    pending = state.get("pending")
    if pending and pending.get("mailer_sent"):
        # The mail was accepted but the run died before the watermark moved. Re-sending
        # would duplicate the escalation, and re-fetching from the old watermark would
        # do the same, so finish that run's bookkeeping here and start today clean.
        banner("Previous run mailed but never committed - closing it out")
        print("Export : {}".format(pathlib.Path(pending["export_file"]).name))
        print("Mailed : yes (recorded before the run stopped) - not sending again")
        commit_run(state, pending["watermark_ticket"], pending["watermark_utc"],
                   pending.get("export_file"), int(pending.get("rows") or 0),
                   mailed=True, modified_utc=pending.get("modified_utc"))
        state = load_state()
        pending = None

    tickets = opt("--tickets")
    if tickets:
        tickets = pathlib.Path(tickets)
        if not tickets.exists():
            raise SystemExit("No such file: {}".format(tickets))
        print("Using existing export: {}".format(tickets.name))
    elif flag("--skip-desk"):
        tickets = newest_export()
        print("Skipping Desk fetch, newest export: {}".format(
            tickets.name if tickets else "none"))
    elif pending and not pending.get("mailer_sent"):
        # A previous run got the tickets but never mailed them. Re-process exactly
        # that data - fetching again would skip it, because the watermark that would
        # have excluded it was never committed.
        tickets = pathlib.Path(pending["export_file"])
        pending["attempts"] = int(pending.get("attempts") or 0) + 1
        save_state(state)
        banner("RESUMING the previous run - its mail never went out")
        print("Export     : {}".format(tickets.name))
        print("Stopped at : {}".format(pending.get("stage_reached")))
        print("Attempt    : {}".format(pending["attempts"]))
        if pending["attempts"] >= 3:
            print("\n*** This run has now failed {} times. If it can never succeed,\n"
                  "*** clear it with:  python main.py --abandon-pending\n"
                  "*** Until then, newer tickets stay queued behind it.".format(
                      pending["attempts"] - 1))
        if not tickets.exists():
            print("\nThe pending export is missing from disk - cannot resume it.")
            print("Clear it with --abandon-pending, or restore the file.")
            return 1
        watermark = {"ticketNumber": pending["watermark_ticket"],
                     "created_utc_iso": pending["watermark_utc"],
                     "modified_utc_iso": pending.get("modified_utc")}
    else:
        banner("PART 1/4  Zoho Desk - fetching logistics tickets")
        result = run_desk_fetch(state, force_today=flag("--today"),
                                force_now=flag("--now"))
        if result is None:
            print("\nNo logistics tickets - nothing to chase in ClickPost.")
            return 0
        tickets = result["path"]
        watermark = {"ticketNumber": result["watermark"]["ticketNumber"],
                     "created_utc_iso": result["watermark"]["created_utc"].isoformat(),
                     "modified_utc_iso": result.get("modified_utc")}

    if tickets is None:
        print("\nNo export available - nothing to do.")
        return 0
    n = awb_count(tickets)
    print("\nExport: {}  ({} AWB(s))".format(tickets.name, n))
    if n == 0:
        print("No AWBs in the export - skipping the ClickPost part.")
        return 0
    if flag("--skip-clickpost"):
        print("\n--skip-clickpost: stopping after part 1.")
        return 0

    # ---- Part 2: ClickPost -------------------------------------------------
    banner("PART 2/4  ClickPost - bulk report for {} AWB(s)".format(n))
    driver = clickpost_session()
    try:
        if not run_bulk_report(driver, tickets):
            print("\nBulk report was not generated - stopping.")
            return 1
        mark_stage(state, "bulk_report")

        banner("PART 2/4  ClickPost - downloading the report")
        csv_path = run_download(driver)
        mark_stage(state, "download")
    except Exception as e:
        print("\nFAILED: {}: {}".format(e.__class__.__name__, e))
        shot(driver, "bulk_error")
        print("Browser left open for inspection.")
        print("The watermark stays put - the next run retries these same tickets.")
        return 1
    finally:
        if flag("--close"):
            try:
                driver.quit()
            except WebDriverException:
                pass

    # ---- Part 3: Merge and Mapping ----------------------------------------
    banner("PART 3/4  Merging status and building Mapping.xlsx")
    try:
        merged = run_merge(tickets, csv_path)
        if merged is None:
            # Every AWB was Delivered/RTO or absent from ClickPost. Nothing to chase
            # and nothing to mail - a finished run, so the watermark may move on.
            print("\nNothing survived the Delivered/RTO filter - no mail to send.")
            if watermark:
                commit_run(state, watermark["ticketNumber"],
                           watermark["created_utc_iso"], tickets, 0, mailed=False,
                           modified_utc=watermark.get("modified_utc_iso"))
                # Committed, so no retry will ever want these files again.
                cleanup_run_files([tickets, csv_path], keep=flag("--keep-files"))
            return 0
        mark_stage(state, "merge")
        print()
        mapping, rows = run_mapping(merged, asof=opt("--asof"))
        mark_stage(state, "mapping")
    except Exception as e:
        print("\nFAILED: {}: {}".format(e.__class__.__name__, e))
        print("The watermark stays put - the next run retries these same tickets.")
        fail_current_run(state, e)
        return 1

    # ---- Part 4: Mail ------------------------------------------------------
    banner("PART 4/4  Mailing {}".format(mapping.name))
    if rows == 0:
        # Asked for explicitly: an empty file is not worth mailing. It still counts
        # as a completed run, or a quiet day would block the pipeline forever.
        print("Mapping file has no rows - no mail sent.")
        mailed = False
    else:
        # Recorded BEFORE cleanup deletes Mapping.xlsx, since the map is read from it.
        columns, ticket_map = thread_ticket_map(mapping)
        message_id = send_mapping_mail(mapping, rows)
        mailed = bool(message_id)
        if not mailed:
            print("\nMail did NOT go out. The watermark stays put, so the next run")
            print("re-processes these same tickets rather than skipping them.")
            fail_current_run(state, "mailer did not accept the message")
            return 1
        # Persisted the moment SMTP accepts, BEFORE anything else can fail. If the
        # process dies between here and commit_run, the next run has to know the mail
        # already went out - otherwise it resumes the pending block and sends the
        # courier a second copy of the same escalation.
        if state.get("pending"):
            state["pending"]["mailer_sent"] = True
            state["pending"]["rows"] = rows
            save_state(state)
        record_thread(message_id, "{} | {}".format(
            MAIL_SUBJECT, datetime.now(IST).strftime(SUBJECT_DATE_FMT)),
            ticket_map, columns)

    names = [tickets.name, csv_path.name, merged.name, mapping.name]
    if watermark:
        commit_run(state, watermark["ticketNumber"], watermark["created_utc_iso"],
                   tickets, rows, mailed=mailed,
                   modified_utc=watermark.get("modified_utc_iso"))
        # The mail is away and the watermark has moved: nothing will be retried, so
        # the day's working files go and tomorrow starts from an empty folder.
        cleanup_run_files([tickets, csv_path, merged, mapping],
                          keep=flag("--keep-files"))
    else:
        print("\n(--tickets/--skip-desk run: watermark and files left untouched)")

    push_to_analytics()

    banner("Done in {:.0f}s".format(time.time() - started))
    print("  tickets  : {}".format(names[0]))
    print("  clickpost: {}".format(names[1]))
    print("  merged   : {}".format(names[2]))
    print("  mapping  : {}  ({} row(s))".format(names[3], rows))
    print("  mailed   : {}".format("yes" if mailed else "no - empty file"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
